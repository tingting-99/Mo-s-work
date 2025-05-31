"""
成绩单和毕业要求跟踪系统 (Transcript and Graduation Requirements Tracker)
版本: 1.0

这个程序用于跟踪高中学生的成绩和毕业要求，支持以下功能:
1. 记录10-12年级的课程成绩
2. 计算GPA (按照AP和CNCC评分标准)
3. 检查毕业要求完成情况
4. 导出成绩单为PDF和Excel格式
5. 保存和加载学生数据

主要模块:
- 数据结构: 定义学生、课程和毕业要求
- GPA计算: 根据不同评分标准计算GPA
- 毕业要求检查: 检查学生是否满足所有毕业要求
- GUI界面: 提供用户友好的界面进行数据管理
- 导出功能: 支持PDF和Excel格式导出
"""

# ================== 导入必要的库 ==================
# 系统和基础库
import sys
import os
import datetime
import json
import pickle
import re
import copy
from difflib import SequenceMatcher

# GUI相关库
import tkinter as tk
from tkinter import messagebox, ttk, filedialog

# 检查Python环境
try:
    # 验证Python版本是否满足要求
    assert sys.version_info >= (3, 6), "Python 3.6 or higher is required"
except Exception as e:
    # 如果Python环境不满足要求，显示错误信息并退出
    error_msg = f"""Python environment error: {str(e)}

Required setup:
1. Install Python 3.6 or higher
2. Create a virtual environment (recommended)
3. Install required packages: pip install -r requirements.txt

Current Python path: {sys.executable}
"""
    messagebox.showerror("Environment Error", error_msg)
    sys.exit(1)

# PDF导出相关库
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch

# Excel导出相关库
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ================== 辅助函数 ==================
def string_similarity(a, b):
    """
    计算两个字符串的相似度

    使用SequenceMatcher计算两个字符串的相似度，返回0到1之间的值，
    1表示完全匹配，0表示完全不匹配。

    参数:
        a (str): 第一个字符串
        b (str): 第二个字符串

    返回:
        float: 相似度，范围为0到1
    """
    # 将字符串转换为小写，并移除多余空格
    a = a.lower().strip()
    b = b.lower().strip()

    # 使用SequenceMatcher计算相似度
    return SequenceMatcher(None, a, b).ratio()

def has_common_words(a, b, min_word_length=3):
    """
    检查两个字符串是否有共同单词

    参数:
        a (str): 第一个字符串
        b (str): 第二个字符串
        min_word_length (int): 最小单词长度，默认为3

    返回:
        tuple: (是否有共同单词, 共同单词列表)
    """
    # 将字符串转换为小写
    a = a.lower()
    b = b.lower()

    # 使用正则表达式提取单词（只保留字母和数字）
    import re
    a_words = set(word for word in re.findall(r'\b[a-z0-9]+\b', a) if len(word) >= min_word_length)
    b_words = set(word for word in re.findall(r'\b[a-z0-9]+\b', b) if len(word) >= min_word_length)

    # 找出共同单词
    common_words = a_words.intersection(b_words)

    return (len(common_words) > 0, list(common_words))

def find_best_match(course_name, course_list, threshold=0.4):
    """
    在课程列表中查找与给定课程名称最相似的课程

    使用多种策略进行匹配：
    1. 首先检查是否有共同单词
    2. 如果有共同单词，计算字符串相似度
    3. 只有当共同单词匹配且相似度超过阈值时，才返回匹配结果

    参数:
        course_name (str): 要匹配的课程名称
        course_list (list): 预设课程名称列表
        threshold (float): 相似度阈值，默认为0.4

    返回:
        tuple: (最佳匹配的课程名称, 相似度, 匹配方式, 共同单词)
    """
    best_match = None
    best_similarity = 0
    best_common_words = []
    match_method = "none"

    for course in course_list:
        # 检查是否有共同单词
        has_common, common_words = has_common_words(course_name, course)

        # 计算字符串相似度
        similarity = string_similarity(course_name, course)

        # 如果有共同单词，增加相似度权重
        if has_common:
            # 根据共同单词的数量和长度增加权重
            word_weight = sum(len(word) for word in common_words) / len(course_name) * 0.5
            adjusted_similarity = similarity + word_weight
        else:
            adjusted_similarity = similarity

        # 更新最佳匹配
        if adjusted_similarity > best_similarity:
            best_similarity = adjusted_similarity
            best_match = course
            best_common_words = common_words
            match_method = "common_words" if has_common else "similarity"

    # 只有当匹配方法是"common_words"且相似度超过阈值时，才返回匹配结果
    if match_method == "common_words" and best_similarity >= threshold:
        return (best_match, best_similarity, match_method, best_common_words)
    else:
        return (None, 0, "none", [])

def get_all_courses():
    """
    从GRADUATION_REQUIREMENTS中提取所有课程名称

    返回:
        list: 所有预设课程名称的列表
    """
    all_courses = []
    for subject_data in GRADUATION_REQUIREMENTS.values():
        all_courses.extend(subject_data['courses'])
    return list(set(all_courses))  # 去除重复项

def round_score(score):
    """
    将分数四舍五入为整数

    参数:
        score (float): 原始分数值

    返回:
        int: 四舍五入后的整数分数
    """
    if score is None:
        return 0

    # 确保分数是数字类型
    try:
        score_float = float(score)
        # 四舍五入为整数
        return int(round(score_float))
    except (ValueError, TypeError):
        return 0

def parse_score(score_value):
    """
    解析分数值，处理各种可能的格式（数字、字符串、百分比等），并四舍五入为整数

    参数:
        score_value: 原始分数值，可能是数字、字符串或None

    返回:
        int: 解析并四舍五入后的整数分数，范围为0-100
    """
    if score_value is None:
        return 0

    # 如果已经是数字类型，检查范围并转换
    if isinstance(score_value, (int, float)):
        score = float(score_value)
        # 如果分数小于1，可能是以小数形式表示的百分比（如0.9333表示93.33%）
        if 0 < score < 1:
            score = score * 100
            print(f"转换小数分数: {score_value} -> {score}")
        # 四舍五入为整数
        return round_score(score)

    # 处理字符串类型
    if isinstance(score_value, str):
        # 移除空白字符
        score_str = score_value.strip()

        # 处理百分比格式 (例如 "86.81%")
        if "%" in score_str:
            # 移除百分号并转换为浮点数
            try:
                score = float(score_str.replace("%", ""))
                return round_score(score)
            except ValueError:
                # 如果转换失败，返回0
                return 0

        # 处理普通数字字符串
        try:
            score = float(score_str)
            # 如果分数小于1，可能是以小数形式表示的百分比（如0.9333表示93.33%）
            if 0 < score < 1:
                score = score * 100
                print(f"转换小数分数字符串: {score_str} -> {score}")
            # 四舍五入为整数
            return round_score(score)
        except ValueError:
            # 如果转换失败，返回0
            return 0

    # 其他情况返回0
    return 0

# ================== 评分标准转换 ==================
"""
定义不同评分标准下的分数到GPA的转换规则
- AP: 美国大学先修课程评分标准
- CNCC: 中国课程评分标准
每个标准包含分数范围到GPA点数的映射
"""
GRADE_SCALE = {
    # AP评分标准 (美国大学先修课程)
    'AP': {
        (97, 100): 4.3,  # A+ (97-100分) = 4.3 GPA
        (90, 96): 4.0,   # A  (90-96分)  = 4.0 GPA
        (80, 89): 3.5,   # B  (80-89分)  = 3.5 GPA
        (70, 79): 3.0,   # C  (70-79分)  = 3.0 GPA
        (60, 69): 2.0,   # D  (60-69分)  = 2.0 GPA
        (0, 59): 0.0     # F  (0-59分)   = 0.0 GPA
    },
    # CNCC评分标准 (中国课程)
    'CNCC': {
        (97, 100): 4.3,  # A+ (97-100分) = 4.3 GPA
        (85, 96): 4.0,   # A  (85-96分)  = 4.0 GPA
        (70, 84): 3.0,   # B  (70-84分)  = 3.0 GPA
        (60, 69): 2.0,   # C  (60-69分)  = 2.0 GPA
        (0, 59): 0.0     # F  (0-59分)   = 0.0 GPA
    }
}

# ================== 毕业要求 ==================
"""
定义各学科的毕业要求
- semesters: 需要完成的学期数
- courses: 该学科类别下的所有课程列表
"""
GRADUATION_REQUIREMENTS = {
    # 中文课程要求 (6个学期)
    'Chinese': {'semesters': 6, 'courses': ['Chinese']},

    # 中国社会科学课程要求 (3个学期)
    'Chinese_Social_Studies': {'semesters': 3, 'courses': ['Chinese History', 'Chinese Geography', 'Chinese Politics']},

    # 英语课程要求 (6个学期)
    'English': {'semesters': 6, 'courses': ['Pre-AP English', 'Pre-AP English Honors', 'AP Seminar 10',
                                           'English 11 Honors', 'English 12 Literature Honors',
                                           'AP English Language and Composition', 'AP English Literature']},

    # 数学课程要求 (6个学期)
    'Mathematics': {'semesters': 6, 'courses': ['Pre-Calculus', 'AP Pre-Calculus', 'AP Calculus AB',
                                               'AP Calculus BC', 'AP Statistics',
                                               'Calculus III',
                                               'Differential Equations', 'Probability Theory']},

    # 社会科学课程要求 (4个学期)
    'Social_Science': {'semesters': 4, 'courses': ['Advanced Economics', 'Developmental Psychology Seminar',
                                                  'Political Science & Law', 'AP Microeconomics',
                                                  'AP Micro/Macroeconomics', 'AP Psychology',
                                                  'AP Environmental Science', 'Intro to Psychology',
                                                  'Sociology', 'Spanish I', 'Spanish II',
                                                  'AP Spanish Language and Culture', 'German I',
                                                  'German II', 'German III', 'French I', 'French II',
                                                  'AP French', 'Japanese I', 'Japanese II',
                                                  'Japanese III', 'Korean', 'Russian']},

    # 科学课程要求 (6个学期)
    'Science': {'semesters': 6, 'courses': ['Biology', 'Biology Honors', 'AP Biology',
                                           'Neuropathology', 'Marine Biology', 'Chemistry Honors',
                                           'AP Chemistry', 'Physics', 'AP Physics C: Mechanics', 'AP Physics C: E/M',
                                           'AP Physics 1','AP Physics 2', 'Advanced Physics']},

    # 技术课程要求 (2个学期) - 不计入GPA
    'Technology': {'semesters': 2, 'courses': ['Technology']},

    # 艺术课程要求 (2个学期) - 不计入GPA
    'Fine_And_Performing_Arts': {'semesters': 2, 'courses': ['AP 3-D Art and Design', 'AP 2-D Art and Design',
                                                            'AP Drawing', 'Visual Arts', 'Visual Arts II',
                                                            'Photography', 'AP Music Theory',
                                                            'Instrumental Ensemble I (Y10 Introduction)',
                                                            'Instrumental Ensemble I (CNCC Semester Introduction)',
                                                            'Instrumental Ensemble II', 'Vocal Ensemble I',
                                                            'Vocal Ensemble II', 'Guitar I', 'Guitar II',
                                                            'Dance', 'Dance I', 'Dance II', 'Drama I',
                                                            'Drama II', 'Independent Video Production']},

    # 体育课程要求 (6个学期) - 不计入GPA
    'Physical_Education': {'semesters': 6, 'courses': ['PE']},

    # 选修课程要求 (6个学期)
    'Electives': {'semesters': 6, 'courses': ['AP Computer Science A', 'AP Computer Science Principles',
                                             'Graphics Programming in Java', 'AP Seminar', 'AP Research',
                                             'Entrepreneurship', 'Web-Development', 'Speech and Debate']},

    # 跨学科研讨课程要求 (1个学期)
    'Interdisciplinary_Seminar': {'semesters': 1, 'courses': ['Interdisciplinary Research Seminar']}
}

# 不计入GPA的学科列表
# 这些学科的课程在计算GPA时会被排除
NON_GPA_SUBJECTS = ['Technology', 'Physical_Education', 'Fine_And_Performing_Arts']

# ================== 数据结构定义 ==================
class Student:
    """
    学生类: 存储学生的基本信息、成绩和毕业要求完成情况

    属性:
        name (str): 学生姓名
        student_id (str): 学生ID
        chinese_name (str): 中文姓名
        date_of_birth (str): 出生日期
        gender (str): 性别
        curriculum_program (str): 课程项目
        date_enrolled (str): 入学日期
        date_graduation (str): 毕业日期
        grades (dict): 学生各年级各学期的成绩，格式为:
                      {年级: {学期: {课程ID: 课程信息}}}
        requirements (dict): 学生毕业要求完成情况，格式为:
                           {学科: {required: 需要学期数, taken: 已完成学期数}}
                           或者对于中国社会科学:
                           {学科: {required_courses: 需要课程数, taken_courses: 已完成课程集合}}
    """
    def __init__(self, name="", student_id="", chinese_name="", date_of_birth="", gender="",
                 curriculum_program="", date_enrolled="", date_graduation=""):
        """
        初始化学生对象

        参数:
            name (str): 学生姓名，默认为空字符串
            student_id (str): 学生ID，默认为空字符串
            chinese_name (str): 中文姓名，默认为空字符串
            date_of_birth (str): 出生日期，默认为空字符串
            gender (str): 性别，默认为空字符串
            curriculum_program (str): 课程项目，默认为空字符串
            date_enrolled (str): 入学日期，默认为空字符串
            date_graduation (str): 毕业日期，默认为空字符串
        """
        # 基本信息
        self.name = name
        self.student_id = student_id
        self.chinese_name = chinese_name
        self.date_of_birth = date_of_birth
        self.gender = gender
        self.curriculum_program = curriculum_program
        self.date_enrolled = date_enrolled
        self.date_graduation = date_graduation

        # 初始化成绩数据结构 - 三个年级(10-12)，每个年级两个学期(Fall/Spring)
        self.grades = {
            "10": {"Fall": {}, "Spring": {}},
            "11": {"Fall": {}, "Spring": {}},
            "12": {"Fall": {}, "Spring": {}}
        }

        # 初始化毕业要求跟踪
        self.requirements = {}

        # 根据全局定义的毕业要求初始化学生的毕业要求完成情况
        for subject, req in GRADUATION_REQUIREMENTS.items():
            # 中国社会科学课程需要特殊处理 - 跟踪具体课程而不仅是学期数
            if subject == 'Chinese_Social_Studies':
                self.requirements[subject] = {"required_courses": 3, "taken_courses": set()}
            else:
                # 其他学科只需跟踪完成的学期数
                self.requirements[subject] = {"required": req['semesters'], "taken": 0}

# ================== GPA计算模块 ==================
def calculate_gpa(grades, specific_grade=None, specific_semester=None):
    """
    计算GPA (Grade Point Average)

    根据提供的成绩数据计算GPA。可以计算总体GPA，特定年级的GPA，或特定年级特定学期的GPA。
    计算时会排除标记为"Not Included"的课程以及属于NON_GPA_SUBJECTS类别的课程。

    参数:
        grades (dict): 成绩数据，格式为 {年级: {学期: {课程ID: 课程信息}}}
        specific_grade (str, optional): 如果提供，只计算该年级的GPA；否则计算总体GPA
        specific_semester (str, optional): 如果提供，只计算该学期的GPA；必须与specific_grade一起使用

    返回:
        float: 计算得到的GPA值，如果没有有效课程则返回0
    """
    # 初始化GPA计算所需变量
    total_points = 0  # 总学分点数
    total_credits = 0  # 总学分数

    # 确定要处理的数据范围
    if specific_grade and specific_semester:
        # 情况1: 只处理特定年级的特定学期
        if specific_grade in grades and specific_semester in grades[specific_grade]:
            semester_data = grades[specific_grade][specific_semester]
            # 处理该学期的所有课程
            for subject, course in semester_data.items():
                # 检查课程是否标记为"不计入GPA"
                scale = course.get("scale", "AP")  # 默认使用AP评分标准
                if scale == "Not Included":
                    continue  # 跳过不计入GPA的课程

                # 检查课程是否属于不计入GPA的学科类别
                subject_found = False
                subject_name = course.get("subject", "")  # 获取原始学科名称

                # 方法1: 直接检查学科名称是否在不计入GPA的列表中
                for non_gpa_subject in NON_GPA_SUBJECTS:
                    if subject_name == non_gpa_subject:  # 直接匹配学科名称
                        subject_found = True
                        break

                    # 方法2: 检查课程名称是否在不计入GPA的学科的课程列表中
                    for course_name in GRADUATION_REQUIREMENTS[non_gpa_subject]['courses']:
                        if course['course'] == course_name:
                            subject_found = True
                            break

                    if subject_found:
                        break

                # 如果课程属于不计入GPA的类别，跳过
                if subject_found:
                    continue

                # 获取课程分数
                score = course["score"]

                # 根据评分标准计算GPA点数
                for (low, high), points in GRADE_SCALE[scale].items():
                    if low <= score <= high:
                        # 累加GPA点数 (GPA点数 × 学分数)
                        total_points += points * course["credits"]
                        # 累加总学分数
                        total_credits += course["credits"]
                        break
        else:
            return 0  # 如果指定的年级或学期不存在，返回0

    elif specific_grade:
        # 情况2: 只处理特定年级的所有学期
        if specific_grade in grades:
            # 遍历该年级的所有学期
            for semester, semester_data in grades[specific_grade].items():
                # 处理该学期的所有课程
                for subject, course in semester_data.items():
                    # 检查课程是否标记为"不计入GPA"
                    scale = course.get("scale", "AP")  # 默认使用AP评分标准
                    if scale == "Not Included":
                        continue  # 跳过不计入GPA的课程

                    # 检查课程是否属于不计入GPA的学科类别
                    subject_found = False
                    subject_name = course.get("subject", "")  # 获取原始学科名称

                    # 方法1: 直接检查学科名称是否在不计入GPA的列表中
                    for non_gpa_subject in NON_GPA_SUBJECTS:
                        if subject_name == non_gpa_subject:  # 直接匹配学科名称
                            subject_found = True
                            break

                        # 方法2: 检查课程名称是否在不计入GPA的学科的课程列表中
                        for course_name in GRADUATION_REQUIREMENTS[non_gpa_subject]['courses']:
                            if course['course'] == course_name:
                                subject_found = True
                                break

                        if subject_found:
                            break

                    # 如果课程属于不计入GPA的类别，跳过
                    if subject_found:
                        continue

                    # 获取课程分数
                    score = course["score"]

                    # 根据评分标准计算GPA点数
                    for (low, high), points in GRADE_SCALE[scale].items():
                        if low <= score <= high:
                            # 累加GPA点数 (GPA点数 × 学分数)
                            total_points += points * course["credits"]
                            # 累加总学分数
                            total_credits += course["credits"]
                            break
        else:
            return 0  # 如果指定的年级不存在，返回0

    else:
        # 情况3: 处理所有年级的所有学期 (计算总体GPA)
        # 遍历所有年级
        for year_data in grades.values():
            # 遍历每个年级的所有学期
            for semester_data in year_data.values():
                # 处理该学期的所有课程
                for subject, course in semester_data.items():
                    # 检查课程是否标记为"不计入GPA"
                    scale = course.get("scale", "AP")  # 默认使用AP评分标准
                    if scale == "Not Included":
                        continue  # 跳过不计入GPA的课程

                    # 检查课程是否属于不计入GPA的学科类别
                    subject_found = False
                    subject_name = course.get("subject", "")  # 获取原始学科名称

                    # 方法1: 直接检查学科名称是否在不计入GPA的列表中
                    for non_gpa_subject in NON_GPA_SUBJECTS:
                        if subject_name == non_gpa_subject:  # 直接匹配学科名称
                            subject_found = True
                            break

                        # 方法2: 检查课程名称是否在不计入GPA的学科的课程列表中
                        for course_name in GRADUATION_REQUIREMENTS[non_gpa_subject]['courses']:
                            if course['course'] == course_name:
                                subject_found = True
                                break

                        if subject_found:
                            break

                    # 如果课程属于不计入GPA的类别，跳过
                    if subject_found:
                        continue

                    # 获取课程分数
                    score = course["score"]

                    # 根据评分标准计算GPA点数
                    for (low, high), points in GRADE_SCALE[scale].items():
                        if low <= score <= high:
                            # 累加GPA点数 (GPA点数 × 学分数)
                            total_points += points * course["credits"]
                            # 累加总学分数
                            total_credits += course["credits"]
                            break

    # 计算并返回GPA (总点数除以总学分)
    # 如果没有有效课程(总学分为0)，则返回0
    return total_points / total_credits if total_credits else 0

# ================== 毕业要求检查 ==================
def check_graduation(student):
    """
    检查学生是否满足所有毕业要求

    遍历学生的毕业要求完成情况，检查每个学科是否满足要求。
    对于中国社会科学课程，检查是否完成了所有必修课程；
    对于其他学科，检查是否完成了足够的学期数。

    参数:
        student (Student): 要检查的学生对象

    返回:
        list: 未满足的毕业要求列表，如果全部满足则返回空列表
    """
    # 存储未满足的毕业要求原因
    failed_reasons = []

    # 检查每个学科的要求
    for subject, req_data in student.requirements.items():
        # 特殊处理中国社会科学课程 - 需要完成特定的课程
        if subject == 'Chinese_Social_Studies':
            # 获取该学科要求的所有课程
            required_courses = set(GRADUATION_REQUIREMENTS[subject]['courses'])

            # 检查学生是否完成了所有必修课程
            if not required_courses.issubset(req_data["taken_courses"]):
                # 计算缺少的课程
                missing = required_courses - req_data["taken_courses"]
                # 添加到未满足要求列表
                failed_reasons.append(
                    f"{subject.replace('_', ' ')}: Missing required courses - {', '.join(missing)}"
                )
        else:
            # 处理其他学科 - 检查完成的学期数是否满足要求
            if req_data["taken"] < req_data["required"]:
                # 添加到未满足要求列表
                failed_reasons.append(
                    f"{subject.replace('_', ' ')}: Need {req_data['required']} semesters, only completed {req_data['taken']}"
                )

    # 返回未满足的毕业要求列表
    return failed_reasons

# ================== GUI界面 ==================
class GradeTracker(tk.Tk):
    """
    成绩单和毕业要求跟踪系统的主GUI类

    继承自tkinter.Tk，创建主应用窗口并管理所有UI组件和功能。
    提供学生信息管理、课程成绩录入、GPA计算、毕业要求检查和数据导出等功能。
    """
    def __init__(self):
        """
        初始化GradeTracker应用

        创建主窗口和所有UI组件，设置布局和事件处理。
        """
        # 初始化基类
        super().__init__()

        # 设置窗口标题和大小
        self.title("Transcript and Graduation Requirements Tracker")
        self.geometry("800x600")

        # 初始化文件路径变量，用于保存/加载数据
        self.current_file_path = None

        # ===== 创建顶部版权信息框架 =====
        # 在窗口顶部右侧显示版权信息
        self.top_frame = ttk.Frame(self)
        self.top_frame.pack(fill="x", padx=10, pady=2)

       
        
       

        # ===== 创建学生信息输入区域 =====
        self.student_info_frame = ttk.LabelFrame(self, text="Student Information")
        self.student_info_frame.pack(fill="x", padx=10, pady=5)

        # 第一行 - 基本信息
        # 中文姓名输入
        ttk.Label(self.student_info_frame, text="Chinese Name:").grid(row=0, column=0, padx=5, pady=5, sticky="w")
        self.student_chinese_name_var = tk.StringVar()  # 存储学生中文姓名的变量
        ttk.Entry(self.student_info_frame, textvariable=self.student_chinese_name_var, width=15).grid(row=0, column=1, padx=5, pady=5, sticky="w")

        # 英文姓名输入
        ttk.Label(self.student_info_frame, text="English Name:").grid(row=0, column=2, padx=5, pady=5, sticky="w")
        self.student_name_var = tk.StringVar()  # 存储学生姓名的变量
        ttk.Entry(self.student_info_frame, textvariable=self.student_name_var, width=20).grid(row=0, column=3, padx=5, pady=5, sticky="w")

        # 学生ID输入
        ttk.Label(self.student_info_frame, text="ID No:").grid(row=0, column=4, padx=5, pady=5, sticky="w")
        self.student_id_var = tk.StringVar()  # 存储学生ID的变量
        ttk.Entry(self.student_info_frame, textvariable=self.student_id_var, width=15).grid(row=0, column=5, padx=5, pady=5, sticky="w")

        # 第二行 - 更多信息
        # 出生日期输入
        ttk.Label(self.student_info_frame, text="Date of Birth:").grid(row=1, column=0, padx=5, pady=5, sticky="w")
        self.student_dob_var = tk.StringVar()  # 存储学生出生日期的变量
        ttk.Entry(self.student_info_frame, textvariable=self.student_dob_var, width=15).grid(row=1, column=1, padx=5, pady=5, sticky="w")

        # 性别输入
        ttk.Label(self.student_info_frame, text="Gender:").grid(row=1, column=2, padx=5, pady=5, sticky="w")
        self.student_gender_var = tk.StringVar()  # 存储学生性别的变量
        gender_combo = ttk.Combobox(self.student_info_frame, textvariable=self.student_gender_var, values=["Male", "Female"], width=10)
        gender_combo.grid(row=1, column=3, padx=5, pady=5, sticky="w")

        # 课程项目输入
        ttk.Label(self.student_info_frame, text="Curriculum:").grid(row=1, column=4, padx=5, pady=5, sticky="w")
        self.student_curriculum_var = tk.StringVar()  # 存储学生课程项目的变量
        ttk.Entry(self.student_info_frame, textvariable=self.student_curriculum_var, width=15).grid(row=1, column=5, padx=5, pady=5, sticky="w")

        # 第三行 - 日期信息
        # 入学日期输入
        ttk.Label(self.student_info_frame, text="Date Enrolled:").grid(row=2, column=0, padx=5, pady=5, sticky="w")
        self.student_enrolled_var = tk.StringVar()  # 存储学生入学日期的变量
        ttk.Entry(self.student_info_frame, textvariable=self.student_enrolled_var, width=15).grid(row=2, column=1, padx=5, pady=5, sticky="w")

        # 毕业日期输入
        ttk.Label(self.student_info_frame, text="Date Graduation:").grid(row=2, column=2, padx=5, pady=5, sticky="w")
        self.student_graduation_var = tk.StringVar()  # 存储学生毕业日期的变量
        ttk.Entry(self.student_info_frame, textvariable=self.student_graduation_var, width=15).grid(row=2, column=3, padx=5, pady=5, sticky="w")

        # ===== 初始化数据对象 =====
        # 创建学生对象
        self.student = Student()
        # 加载课程数据库
        self.course_db = self.load_courses()

        # ===== 创建年级选择标签页 =====
        # 使用Notebook组件创建10-12年级的标签页
        self.notebook = ttk.Notebook(self)
        # 为每个年级创建一个标签页
        for grade in ["10", "11", "12"]:
            # 创建年级框架并添加到标签页
            frame = GradeFrame(self.notebook, grade, self.course_db)
            self.notebook.add(frame, text=f"Grade {grade}")
        # 将标签页添加到主窗口
        self.notebook.pack(expand=True, fill="both")

        # ===== 创建控制按钮区域 =====
        # 第一行 - 主要功能控制
        self.control_frame1 = ttk.LabelFrame(self, text="Main Controls")
        self.control_frame1.pack(fill="x", padx=10, pady=5)

        # GPA计算按钮
        self.calculate_btn = ttk.Button(
            self.control_frame1,
            text="Calculate All GPAs",
            command=self.calculate_gpa  # 点击时调用GPA计算方法
        )
        self.calculate_btn.pack(side="left", padx=5, pady=5)

        # 毕业要求检查按钮
        self.check_btn = ttk.Button(
            self.control_frame1,
            text="Check Graduation Requirements",
            command=self.check_graduation_req  # 点击时调用毕业要求检查方法
        )
        self.check_btn.pack(side="left", padx=5, pady=5)

        # 第二行 - 导出选项
        self.control_frame2 = ttk.LabelFrame(self, text="Export Options")
        self.control_frame2.pack(fill="x", padx=10, pady=5)

        # PDF导出按钮
        self.export_pdf_btn = ttk.Button(
            self.control_frame2,
            text="Export PDF Transcript",
            command=self.export_pdf  # 点击时调用PDF导出方法
        )
        self.export_pdf_btn.pack(side="left", padx=5, pady=5)

        # Excel导出按钮
        self.export_excel_btn = ttk.Button(
            self.control_frame2,
            text="Export Excel Transcript",
            command=self.export_excel  # 点击时调用Excel导出方法
        )
        self.export_excel_btn.pack(side="left", padx=5, pady=5)

        # School Transcript PDF导出按钮
        self.export_school_transcript_btn = ttk.Button(
            self.control_frame2,
            text="Export School Transcript",
            command=self.export_school_transcript  # 点击时调用School Transcript导出方法
        )
        self.export_school_transcript_btn.pack(side="left", padx=5, pady=5)

        # 第三行 - 文件操作
        self.control_frame3 = ttk.LabelFrame(self, text="File Operations")
        self.control_frame3.pack(fill="x", padx=10, pady=5)

        # 保存数据按钮
        self.save_btn = ttk.Button(
            self.control_frame3,
            text="Save Data",
            command=self.save_data  # 点击时调用保存数据方法
        )
        self.save_btn.pack(side="left", padx=5, pady=5)

        # 另存为按钮
        self.save_as_btn = ttk.Button(
            self.control_frame3,
            text="Save As...",
            command=self.save_data_as  # 点击时调用另存为方法
        )

        # 导入Excel按钮
        self.import_excel_btn = ttk.Button(
            self.control_frame3,
            text="Import Excel Data",
            command=self.import_excel_data  # 点击时调用导入Excel数据方法
        )
        self.import_excel_btn.pack(side="left", padx=5, pady=5)
        self.save_as_btn.pack(side="left", padx=5, pady=5)

        # 加载数据按钮
        self.load_btn = ttk.Button(
            self.control_frame3,
            text="Load Data",
            command=self.load_data  # 点击时调用加载数据方法
        )
        self.load_btn.pack(side="left", padx=5, pady=5)

        # 合并TGRT文件按钮
        self.merge_tgrt_btn = ttk.Button(
            self.control_frame3,
            text="Merge TGRT Files",
            command=self.merge_tgrt_files  # 点击时调用合并TGRT文件方法
        )
        self.merge_tgrt_btn.pack(side="left", padx=5, pady=5)

        # ===== 创建结果显示区域 =====
        # 用于显示GPA计算结果和毕业要求检查结果
        self.result_frame = ttk.LabelFrame(self, text="Results")
        self.result_frame.pack(fill="both", expand=True, padx=10, pady=5)

        # 文本显示区域
        self.result_text = tk.Text(self.result_frame, height=10)
        self.result_text.pack(fill="both", expand=True, padx=5, pady=5)

        # ===== 底部版权信息 =====
       
        

    def load_courses(self):
        """
        加载课程数据库

        从全局定义的毕业要求中提取所有课程，按学科分类并排序。

        返回:
            dict: 按学科分类的课程字典，格式为 {学科: [课程列表]}
        """
        # 创建按学科分类的课程字典
        courses_by_subject = {}

        # 从GRADUATION_REQUIREMENTS中提取每个学科的课程列表
        for subject, data in GRADUATION_REQUIREMENTS.items():
            # 对每个学科的课程列表进行排序，便于在下拉菜单中显示
            courses_by_subject[subject] = sorted(data['courses'])

        return courses_by_subject

    def calculate_gpa(self):
        """
        计算并显示各学期、各年级和总体GPA

        1. 收集所有课程数据
        2. 计算每个学期的GPA
        3. 计算每个年级的GPA
        4. 计算总体GPA
        5. 在结果区域显示所有GPA数据
        """
        # ===== 1. 收集所有课程数据 =====
        all_grades = {}
        for grade in ["10", "11", "12"]:
            all_grades[grade] = {}
            for semester in ["Fall", "Spring"]:
                # 获取该学期的所有课程成绩
                all_grades[grade][semester] = self.get_semester_grades(grade, semester)

        # ===== 2. 计算每个学期的GPA =====
        semester_gpas = {}
        for grade in ["10", "11", "12"]:
            semester_gpas[grade] = {}
            for semester in ["Fall", "Spring"]:
                # 调用GPA计算函数，只计算特定年级特定学期的GPA
                semester_gpas[grade][semester] = calculate_gpa(all_grades, grade, semester)

        # ===== 3. 计算每个年级的GPA =====
        grade_gpas = {}
        for grade in ["10", "11", "12"]:
            # 调用GPA计算函数，计算特定年级的GPA (包括该年级的所有学期)
            grade_gpas[grade] = calculate_gpa(all_grades, grade)

        # ===== 4. 计算总体GPA =====
        # 调用GPA计算函数，计算所有年级所有学期的总体GPA
        overall_gpa = calculate_gpa(all_grades)

        # ===== 5. 在结果区域显示所有GPA数据 =====
        # 清空结果文本区域
        self.result_text.delete(1.0, tk.END)

        # 显示各年级和学期的GPA
        self.result_text.insert(tk.END, "===== GPA by Year and Semester =====\n")

        # 遍历每个年级
        for grade in ["10", "11", "12"]:
            # 计算该年级的课程总数
            course_count = 0
            for semester in ["Fall", "Spring"]:
                course_count += len(all_grades[grade][semester])

            # 如果该年级有课程数据，显示该年级的GPA
            if course_count > 0:
                self.result_text.insert(tk.END, f"\nGrade {grade} Year GPA: {grade_gpas[grade]:.2f}\n")

                # 显示该年级每个学期的GPA
                for semester in ["Fall", "Spring"]:
                    semester_course_count = len(all_grades[grade][semester])
                    if semester_course_count > 0:
                        # 显示学期GPA和课程数量
                        self.result_text.insert(tk.END,
                            f"  - {semester} Semester GPA: {semester_gpas[grade][semester]:.2f} ({semester_course_count} courses)\n")
                    else:
                        # 如果该学期没有课程数据
                        self.result_text.insert(tk.END, f"  - {semester} Semester GPA: No courses data\n")
            else:
                # 如果该年级没有课程数据
                self.result_text.insert(tk.END, f"\nGrade {grade} Year GPA: No courses data\n")

        # 显示总体GPA
        self.result_text.insert(tk.END, "\n===== Overall GPA =====\n")

        # 计算所有课程总数
        total_course_count = 0
        for grade in ["10", "11", "12"]:
            for semester in ["Fall", "Spring"]:
                total_course_count += len(all_grades[grade][semester])

        # 显示总体GPA和课程总数
        if total_course_count > 0:
            self.result_text.insert(tk.END, f"Overall GPA: {overall_gpa:.2f} (Total: {total_course_count} courses)\n")
        else:
            # 如果没有任何课程数据
            self.result_text.insert(tk.END, "Overall GPA: No courses data\n")

    def get_semester_grades(self, grade, semester):
        """
        获取指定年级和学期的所有课程成绩

        从UI界面中收集指定年级和学期的所有课程数据，并转换为适合GPA计算的格式。
        为每个课程创建唯一的键，即使它们属于相同的学科。

        参数:
            grade (str): 年级 ("10", "11", "12")
            semester (str): 学期 ("Fall", "Spring")

        返回:
            dict: 课程成绩字典，格式为 {课程键: 课程数据}
        """
        # 存储所有课程成绩的字典
        grades = {}
        # 用于为同一学科的多个课程创建唯一键的计数器
        course_counter = {}

        # 在标签页中查找对应的年级框架
        for i in range(self.notebook.index("end")):
            frame = self.notebook.winfo_children()[i]
            # 检查是否是目标年级的框架
            if isinstance(frame, GradeFrame) and frame.grade == grade:
                # 在年级框架中查找对应的学期框架
                for sem_frame in frame.semester_notebook.winfo_children():
                    # 检查是否是目标学期的框架
                    if isinstance(sem_frame, SemesterFrame) and sem_frame.semester == semester:
                        # 获取该学期的所有课程数据
                        for entry in sem_frame.get_courses():
                            # 解析课程数据
                            subject, course, score, scale = entry

                            # 为每个课程创建唯一的键，即使它们属于相同的学科
                            if subject in course_counter:
                                # 如果该学科已有课程，递增计数器
                                course_counter[subject] += 1
                                key = f"{subject}_{course_counter[subject]}"
                            else:
                                # 如果是该学科的第一个课程
                                course_counter[subject] = 1
                                key = f"{subject}_1"

                            # 使用唯一键存储课程数据
                            grades[key] = {
                                "subject": subject,  # 存储原始学科名称，用于后续判断是否计入GPA
                                "course": course,    # 课程名称
                                "score": score,      # 课程分数
                                "scale": scale,      # 评分标准 (AP/CNCC/Not Included)
                                "credits": 1         # 学分数 (默认为1)
                            }
        # 返回收集到的所有课程成绩
        return grades

    def check_graduation_req(self):
        """
        检查并显示毕业要求完成状态

        1. 更新学生数据
        2. 检查毕业要求完成情况
        3. 在结果区域显示检查结果

        返回:
            list: 未满足的毕业要求列表，如果全部满足则返回空列表
        """
        # 1. 更新学生数据 - 确保使用最新的课程数据
        self.update_student_data()

        # 2. 检查毕业要求完成情况
        failed_reasons = check_graduation(self.student)

        # 3. 在结果区域显示检查结果
        # 清空结果文本区域
        self.result_text.delete(1.0, tk.END)

        # 根据检查结果显示不同的信息
        if failed_reasons:
            # 如果有未满足的要求，显示未满足的要求列表
            self.result_text.insert(tk.END, "Graduation Requirements Not Met:\n")

            # 遍历每个未满足的要求，并以红色显示
            for reason in failed_reasons:
                # 使用标签控制文本颜色 - 项目符号使用正常颜色
                self.result_text.insert(tk.END, "- ", "normal")
                # 未满足的要求使用红色显示
                self.result_text.insert(tk.END, reason, "failed")
                # 换行使用正常颜色
                self.result_text.insert(tk.END, "\n", "normal")

            # 配置"failed"标签为红色文本
            self.result_text.tag_configure("failed", foreground="red")
        else:
            # 如果所有要求都满足，显示祝贺信息
            self.result_text.insert(tk.END, "Congratulations! All graduation requirements have been met.")

        # 返回未满足的要求列表，供其他方法使用
        return failed_reasons

    def save_data(self):
        """
        保存学生数据到当前文件或提示选择新文件

        如果已有当前文件路径，则直接保存到该文件；
        否则调用save_data_as方法提示用户选择保存位置。
        """
        # 检查是否已有保存文件路径
        if not self.current_file_path:
            # 如果没有，调用另存为方法
            self.save_data_as()
        else:
            # 如果有，直接保存到当前文件
            self._save_to_file(self.current_file_path)

    def save_data_as(self):
        """
        将学生数据保存到新文件

        打开文件保存对话框，让用户选择保存位置和文件名，
        然后将数据保存到选定的文件。
        """
        # 打开文件保存对话框
        file_path = filedialog.asksaveasfilename(
            defaultextension=".tgrt",  # 默认文件扩展名
            filetypes=[("Transcript Data Files", "*.tgrt"), ("All Files", "*.*")],  # 文件类型过滤
            title="Save Transcript Data"  # 对话框标题
        )

        # 检查用户是否取消了操作
        if not file_path:
            return  # 用户取消，直接返回

        # 保存数据到选定的文件
        self._save_to_file(file_path)
        # 更新当前文件路径
        self.current_file_path = file_path

    def _save_to_file(self, file_path):
        """
        将数据保存到指定文件

        1. 更新学生数据
        2. 收集所有课程数据
        3. 将数据序列化并保存到文件

        参数:
            file_path (str): 要保存到的文件路径
        """
        try:
            # 1. 更新学生数据，确保使用最新的信息
            self.update_student_data()

            # 2. 收集所有课程数据，创建要保存的数据结构
            data = {
                # 学生基本信息
                "student": {
                    "name": self.student.name,
                    "student_id": self.student.student_id,
                    "chinese_name": self.student.chinese_name,
                    "date_of_birth": self.student.date_of_birth,
                    "gender": self.student.gender,
                    "curriculum_program": self.student.curriculum_program,
                    "date_enrolled": self.student.date_enrolled,
                    "date_graduation": self.student.date_graduation
                },
                # 课程数据
                "courses": {}
            }

            # 从UI中获取每个年级和学期的课程数据
            for grade in ["10", "11", "12"]:
                data["courses"][grade] = {}
                for semester in ["Fall", "Spring"]:
                    courses = []

                    # 在标签页中查找对应的年级框架
                    for i in range(self.notebook.index("end")):
                        frame = self.notebook.winfo_children()[i]
                        if isinstance(frame, GradeFrame) and frame.grade == grade:
                            # 在年级框架中查找对应的学期框架
                            for sem_frame in frame.semester_notebook.winfo_children():
                                if isinstance(sem_frame, SemesterFrame) and sem_frame.semester == semester:
                                    # 获取该学期的所有课程数据并转换为字典格式
                                    courses = [
                                        {
                                            "subject": entry[0],  # 学科
                                            "course": entry[1],   # 课程名称
                                            "score": entry[2],    # 分数
                                            "scale": entry[3]     # 评分标准
                                        }
                                        for entry in sem_frame.get_courses()
                                    ]

                    # 将该学期的课程数据添加到数据结构中
                    data["courses"][grade][semester] = courses

            # 3. 将数据序列化并保存到文件
            with open(file_path, 'wb') as f:
                pickle.dump(data, f)  # 使用pickle序列化数据

            # 显示保存成功消息
            messagebox.showinfo("Success", f"Data saved to {file_path}")

        except Exception as e:
            # 如果保存过程中出现错误，显示错误消息
            messagebox.showerror("Error", f"Failed to save data: {str(e)}")

    def fix_decimal_scores(self, data):
        """
        修复数据中的小数分数问题并应用四舍五入

        1. 检查所有课程的分数，如果发现小于1的分数，将其乘以100
        2. 将所有分数四舍五入为整数

        参数:
            data: 包含课程数据的字典

        返回:
            tuple: (修复后的数据, 修复的分数计数, 四舍五入的分数计数)
        """
        # 创建数据的深拷贝，避免修改原始数据
        fixed_data = copy.deepcopy(data)
        decimal_fixed_count = 0
        rounded_count = 0

        # 遍历所有年级和学期
        for grade in ["10", "11", "12"]:
            if grade in fixed_data["courses"]:
                for semester in ["Fall", "Spring"]:
                    if semester in fixed_data["courses"][grade]:
                        courses = fixed_data["courses"][grade][semester]
                        for i, course_info in enumerate(courses):
                            score = course_info["score"]
                            original_score = score

                            # 检查分数是否为小数（0-1范围内）
                            if isinstance(score, (int, float)) and 0 < score < 1:
                                # 将分数乘以100
                                score = score * 100
                                print(f"修复小数分数: {course_info['course']} - {original_score} -> {score}")
                                decimal_fixed_count += 1

                            # 应用四舍五入
                            if isinstance(score, (int, float)) and not isinstance(score, int):
                                rounded_score = round_score(score)
                                if rounded_score != score:
                                    print(f"四舍五入分数: {course_info['course']} - {score} -> {rounded_score}")
                                    rounded_count += 1
                                score = rounded_score

                            # 更新分数
                            fixed_data["courses"][grade][semester][i]["score"] = score

        if decimal_fixed_count > 0:
            print(f"共修复了 {decimal_fixed_count} 个小数分数")
        if rounded_count > 0:
            print(f"共四舍五入了 {rounded_count} 个分数")

        return fixed_data, decimal_fixed_count, rounded_count

    def load_data(self):
        """
        从文件加载学生数据

        1. 打开文件选择对话框
        2. 从选定的文件加载数据
        3. 更新UI界面显示加载的数据
        4. 计算并显示GPA
        """
        # 1. 打开文件选择对话框
        file_path = filedialog.askopenfilename(
            filetypes=[("Transcript Data Files", "*.tgrt"), ("All Files", "*.*")],  # 文件类型过滤
            title="Load Transcript Data"  # 对话框标题
        )

        # 检查用户是否取消了操作
        if not file_path:
            return  # 用户取消，直接返回

        try:
            # 2. 从选定的文件加载数据
            with open(file_path, 'rb') as f:
                data = pickle.load(f)  # 使用pickle反序列化数据

            # 修复可能的小数分数问题并应用四舍五入
            fixed_data, decimal_fixed_count, rounded_count = self.fix_decimal_scores(data)
            data = fixed_data

            # 如果修复了分数或应用了四舍五入，询问用户是否要保存修复后的数据
            if decimal_fixed_count > 0 or rounded_count > 0:
                message = ""
                if decimal_fixed_count > 0:
                    message += f"Found and fixed {decimal_fixed_count} decimal scores (e.g., 0.93 → 93).\n"
                if rounded_count > 0:
                    message += f"Rounded {rounded_count} scores to integers (e.g., 93.5 → 94).\n"
                message += "Would you like to save the updated data?"

                save_fixed = messagebox.askyesno(
                    "Fix and Round Scores",
                    message
                )
                if save_fixed:
                    # 保存修复后的数据
                    with open(file_path, 'wb') as f:
                        pickle.dump(data, f)
                    messagebox.showinfo("Success", f"Updated data saved to {file_path}")

            # 3. 更新UI界面显示加载的数据
            # 3.1 更新学生基本信息
            if "student" in data:
                # 设置学生所有信息输入框的值
                self.student_name_var.set(data["student"].get("name", ""))
                self.student_id_var.set(data["student"].get("student_id", ""))
                self.student_chinese_name_var.set(data["student"].get("chinese_name", ""))
                self.student_dob_var.set(data["student"].get("date_of_birth", ""))
                self.student_gender_var.set(data["student"].get("gender", ""))
                self.student_curriculum_var.set(data["student"].get("curriculum_program", ""))
                self.student_enrolled_var.set(data["student"].get("date_enrolled", ""))
                self.student_graduation_var.set(data["student"].get("date_graduation", ""))

            # 3.2 清除现有的所有课程数据
            for i in range(self.notebook.index("end")):
                frame = self.notebook.winfo_children()[i]
                if isinstance(frame, GradeFrame):
                    for sem_frame in frame.semester_notebook.winfo_children():
                        if isinstance(sem_frame, SemesterFrame):
                            # 清空课程列表
                            for item_id in sem_frame.courses_list.get_children():
                                sem_frame.courses_list.delete(item_id)
                            # 重置课程条目列表
                            sem_frame.course_entries = []

            # 3.3 添加从文件加载的课程数据
            if "courses" in data:
                for grade, grade_data in data["courses"].items():
                    for semester, courses in grade_data.items():
                        # 在标签页中查找对应的年级框架
                        for i in range(self.notebook.index("end")):
                            frame = self.notebook.winfo_children()[i]
                            if isinstance(frame, GradeFrame) and frame.grade == grade:
                                # 在年级框架中查找对应的学期框架
                                for sem_frame in frame.semester_notebook.winfo_children():
                                    if isinstance(sem_frame, SemesterFrame) and sem_frame.semester == semester:
                                        # 添加课程数据到UI
                                        for course_data in courses:
                                            # 提取课程信息
                                            subject = course_data.get("subject", "")  # 学科
                                            course = course_data.get("course", "")    # 课程名称
                                            score = course_data.get("score", 0)       # 分数
                                            scale = course_data.get("scale", "AP")    # 评分标准

                                            # 添加到课程列表
                                            item_id = sem_frame.courses_list.insert(
                                                "", "end",
                                                values=(subject, course, score, scale)
                                            )
                                            # 添加到课程条目列表
                                            sem_frame.course_entries.append(
                                                (subject, course, score, scale, item_id)
                                            )

            # 更新当前文件路径
            self.current_file_path = file_path

            # 4. 计算并显示GPA
            self.calculate_gpa()

            # 显示加载成功消息
            messagebox.showinfo("Success", f"Data loaded from {file_path}")

        except Exception as e:
            # 如果加载过程中出现错误，显示错误消息
            messagebox.showerror("Error", f"Failed to load data: {str(e)}")

    def export_excel(self):
        """Export Excel transcript with the same content as PDF"""
        # Ensure data is up-to-date
        self.calculate_gpa()
        failed_reasons = self.check_graduation_req()

        # Collect all course data
        all_grades = {}
        for grade in ["10", "11", "12"]:
            all_grades[grade] = {}
            for semester in ["Fall", "Spring"]:
                all_grades[grade][semester] = self.get_semester_grades(grade, semester)

        # Calculate GPA for each semester
        semester_gpas = {}
        for grade in ["10", "11", "12"]:
            semester_gpas[grade] = {}
            for semester in ["Fall", "Spring"]:
                semester_gpas[grade][semester] = calculate_gpa(all_grades, grade, semester)

        # Calculate GPA for each year
        grade_gpas = {}
        for grade in ["10", "11", "12"]:
            grade_gpas[grade] = calculate_gpa(all_grades, grade)

        # Calculate overall GPA
        overall_gpa = calculate_gpa(all_grades)

        # Choose save location
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")],
            title="Save Transcript Excel"
        )

        if not file_path:
            return  # User canceled

        try:
            # Create Excel workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Transcript"

            # Define styles - 调整字体大小
            title_font = Font(name='Arial', size=14, bold=True)  # 减小标题字体
            subtitle_font = Font(name='Arial', size=12, bold=True)  # 减小副标题字体
            normal_font = Font(name='Arial', size=9)  # 减小正文字体

            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

            center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)  # 添加自动换行

            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Add title
            ws['A1'] = "Student Transcript"
            ws['A1'].font = title_font
            ws.merge_cells('A1:E1')
            ws['A1'].alignment = center_align

            # Add student information table - 所有信息放在一行
            row = 2

            # 添加表头 - 包含所有字段
            headers = ["Name in Chinese", "Name in English", "ID No", "Date of Birth", "Gender", "Curriculum Program", "Date Enrolled", "Date Graduation"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = header
                cell.font = Font(name='Arial', size=8, bold=True)  # 进一步减小字体大小
                cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")  # Light grey
                cell.alignment = center_align
                cell.border = thin_border
            row += 1

            # 添加学生信息 - 所有信息在一行
            student_info = [
                self.student.chinese_name or "",
                self.student.name or "",
                self.student.student_id or "",
                self.student.date_of_birth or "",
                self.student.gender or "",
                self.student.curriculum_program or "",
                self.student.date_enrolled or "",
                self.student.date_graduation or ""
            ]
            for col, info in enumerate(student_info, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = info
                cell.font = Font(name='Arial', size=8)  # 进一步减小字体大小
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = thin_border

            row += 2  # Add space

            # Add overall GPA
            ws[f'A{row}'] = f"Overall GPA: {overall_gpa:.2f}"
            ws[f'A{row}'].font = subtitle_font
            ws.merge_cells(f'A{row}:E{row}')
            row += 2  # Add space

            # Add each year's GPA and courses
            for grade in ["10", "11", "12"]:
                # Add year header and GPA
                if grade_gpas.get(grade, 0) > 0:
                    ws[f'A{row}'] = f"Grade {grade}"
                    ws[f'A{row}'].font = subtitle_font
                    ws.merge_cells(f'A{row}:E{row}')
                    row += 1

                    ws[f'A{row}'] = f"Year GPA: {grade_gpas[grade]:.2f}"
                    ws[f'A{row}'].font = normal_font
                    ws.merge_cells(f'A{row}:E{row}')
                    row += 1

                    # Add semester GPAs
                    for semester in ["Fall", "Spring"]:
                        if semester_gpas[grade].get(semester, 0) > 0:
                            ws[f'A{row}'] = f"{semester} Semester GPA: {semester_gpas[grade][semester]:.2f}"
                            ws[f'A{row}'].font = normal_font
                            ws.merge_cells(f'A{row}:E{row}')
                            row += 1

                    # Add course table headers - 进一步减小字体大小
                    headers = ['Semester', 'Course', 'Score', 'Scale', 'GPA']
                    for col, header in enumerate(headers, 1):
                        cell = ws.cell(row=row, column=col)
                        cell.value = header
                        cell.font = Font(name='Arial', size=8, bold=True, color="FFFFFF")  # 进一步减小字体大小
                        cell.fill = header_fill
                        cell.alignment = center_align
                        cell.border = thin_border
                    row += 1

                    # Add courses from both semesters
                    for semester in ["Fall", "Spring"]:
                        courses = all_grades[grade][semester]

                        for _, course_data in courses.items():
                            # Calculate individual course GPA
                            course_gpa = "N/A"  # Default value

                            # Check if course counts for GPA
                            scale = course_data.get("scale", "AP")

                            # Skip courses marked as "Not Included"
                            if scale == "Not Included":
                                course_gpa = "N/A"
                            else:
                                # Check if it's a non-GPA subject based on subject category
                                subject_found = False
                                subject_name = course_data.get("subject", "")
                                for non_gpa_subject in NON_GPA_SUBJECTS:
                                    if subject_name == non_gpa_subject:
                                        subject_found = True
                                        break
                                    for course_name in GRADUATION_REQUIREMENTS[non_gpa_subject]['courses']:
                                        if course_data['course'] == course_name:
                                            subject_found = True
                                            break
                                    if subject_found:
                                        break

                                if not subject_found:
                                    score = course_data["score"]

                                    # Calculate GPA
                                    for (low, high), points in GRADE_SCALE[scale].items():
                                        if low <= score <= high:
                                            course_gpa = f"{points:.1f}"
                                            break

                            # Add course data to Excel - 减小字体大小
                            values = [
                                semester,
                                course_data['course'],
                                str(course_data['score']),
                                course_data.get('scale', 'AP'),
                                course_gpa
                            ]

                            for col, value in enumerate(values, 1):
                                cell = ws.cell(row=row, column=col)
                                cell.value = value
                                cell.font = Font(name='Arial', size=8)  # 减小字体大小
                                cell.border = thin_border
                                if col == 1:  # Semester column
                                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                                elif col == 2:  # Course column
                                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                                else:
                                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                            row += 1

                    row += 1  # Add space between grades

            # Add graduation requirements section
            row += 1
            ws[f'A{row}'] = "Graduation Requirements Status"
            ws[f'A{row}'].font = subtitle_font
            ws.merge_cells(f'A{row}:E{row}')
            row += 1

            if failed_reasons:
                ws[f'A{row}'] = "Graduation Requirements Not Met:"
                ws[f'A{row}'].font = normal_font
                ws.merge_cells(f'A{row}:E{row}')
                row += 1

                for reason in failed_reasons:
                    ws[f'A{row}'] = f"- {reason}"
                    ws[f'A{row}'].font = normal_font
                    ws[f'A{row}'].font = Font(name='Arial', size=11, color="FF0000")  # Red color for failed requirements
                    ws.merge_cells(f'A{row}:E{row}')
                    row += 1
            else:
                ws[f'A{row}'] = "Congratulations! All graduation requirements have been met."
                ws[f'A{row}'].font = normal_font
                ws.merge_cells(f'A{row}:E{row}')
                row += 1

            # Adjust column widths - 优化列宽以适应内容
            ws.column_dimensions['A'].width = 12  # Chinese Name
            ws.column_dimensions['B'].width = 12  # English Name
            ws.column_dimensions['C'].width = 10  # ID No
            ws.column_dimensions['D'].width = 10  # Date of Birth
            ws.column_dimensions['E'].width = 8   # Gender
            ws.column_dimensions['F'].width = 12  # Curriculum Program
            ws.column_dimensions['G'].width = 10  # Date Enrolled
            ws.column_dimensions['H'].width = 10  # Date Graduation

            # 设置行高，确保内容不会被截断
            for i in range(1, row+1):
                ws.row_dimensions[i].height = 20

            

            # Save the workbook
            wb.save(file_path)
            messagebox.showinfo("Success", f"Excel transcript saved to {file_path}")

        except Exception as e:
            messagebox.showerror("Error", f"Failed to create Excel file: {str(e)}")

    def import_excel_data(self):
        """
        从Excel文件导入学生数据并为每个学生生成.tgrt文件

        1. 打开文件选择对话框，让用户选择Excel文件
        2. 打开对话框让用户选择学期（如G10 Fall）
        3. 解析Excel文件，提取学生和课程数据
        4. 为每个学生创建一个Student对象，并填充课程数据
        5. 为每个学生生成一个.tgrt文件
        6. 显示处理结果
        """
        # 调用import_excel方法
        self.import_excel()

    def merge_tgrt_files(self):
        """
        合并多个tgrt文件的功能

        1. 允许用户选择最多6个tgrt文件，每个对应一个学期
        2. 检查所有文件是否属于同一个学生
        3. 合并所有文件中的课程数据
        4. 保存为一个新的tgrt文件
        """
        try:
            # 创建学期选择对话框
            merge_dialog = tk.Toplevel(self)
            merge_dialog.title("Merge TGRT Files")
            merge_dialog.geometry("500x500")
            merge_dialog.resizable(False, False)
            merge_dialog.transient(self)
            merge_dialog.grab_set()

            # 创建对话框内容
            frame = ttk.Frame(merge_dialog, padding=10)
            frame.pack(fill="both", expand=True)

            # 添加说明标签
            ttk.Label(
                frame,
                text="Select TGRT files for each semester you want to merge.\nAll files must belong to the same student.",
                wraplength=480
            ).pack(pady=10)

            # 创建学期文件选择框架
            semester_files_frame = ttk.Frame(frame)
            semester_files_frame.pack(fill="x", pady=5)

            # 定义学期列表
            semesters = [
                "G10 Fall", "G10 Spring",
                "G11 Fall", "G11 Spring",
                "G12 Fall", "G12 Spring"
            ]

            # 存储每个学期对应的文件路径
            semester_file_paths = {semester: None for semester in semesters}

            # 为每个学期创建文件选择控件
            semester_file_labels = {}

            # 创建一个框架来包含所有学期的文件选择控件
            for i, semester in enumerate(semesters):
                # 创建学期框架
                semester_frame = ttk.Frame(semester_files_frame)
                semester_frame.pack(fill="x", pady=2)

                # 添加学期标签
                ttk.Label(semester_frame, text=f"{semester}:", width=10).pack(side="left")

                # 添加文件路径标签
                file_label = ttk.Label(semester_frame, text="No file selected", width=40)
                file_label.pack(side="left", padx=5)
                semester_file_labels[semester] = file_label

                # 创建文件选择按钮
                def create_file_selector(sem, label):
                    def select_file():
                        file_path = filedialog.askopenfilename(
                            title=f"Select TGRT File for {sem}",
                            filetypes=[("TGRT Files", "*.tgrt")],
                            initialdir=os.getcwd()
                        )
                        if file_path:
                            semester_file_paths[sem] = file_path
                            # 显示文件名而不是完整路径
                            file_name = os.path.basename(file_path)
                            if len(file_name) > 35:
                                file_name = file_name[:32] + "..."
                            label.config(text=file_name)
                    return select_file

                file_btn = ttk.Button(
                    semester_frame,
                    text="Browse",
                    command=create_file_selector(semester, file_label),
                    width=10
                )
                file_btn.pack(side="right", padx=5)

            # 添加分隔线
            ttk.Separator(frame, orient="horizontal").pack(fill="x", pady=10)

            # 添加输出文件选择区域
            output_frame = ttk.Frame(frame)
            output_frame.pack(fill="x", pady=5)

            ttk.Label(output_frame, text="Output File:").pack(side="left")

            output_file_var = tk.StringVar(value="")
            output_file_label = ttk.Label(output_frame, text="No file selected", width=40)
            output_file_label.pack(side="left", padx=5)

            # 输出文件选择函数
            def select_output_file():
                file_path = filedialog.asksaveasfilename(
                    title="Save Merged TGRT File",
                    filetypes=[("TGRT Files", "*.tgrt")],
                    defaultextension=".tgrt",
                    initialdir=os.getcwd()
                )
                if file_path:
                    output_file_var.set(file_path)
                    # 显示文件名而不是完整路径
                    file_name = os.path.basename(file_path)
                    if len(file_name) > 35:
                        file_name = file_name[:32] + "..."
                    output_file_label.config(text=file_name)

            output_file_btn = ttk.Button(
                output_frame,
                text="Browse",
                command=select_output_file,
                width=10
            )
            output_file_btn.pack(side="right", padx=5)

            # 添加分隔线
            ttk.Separator(frame, orient="horizontal").pack(fill="x", pady=10)

            # 添加合并按钮和取消按钮
            btn_frame = ttk.Frame(frame)
            btn_frame.pack(fill="x", pady=10)

            # 合并按钮回调函数
            def on_merge():
                # 获取所有选择的文件路径
                selected_files = {sem: path for sem, path in semester_file_paths.items() if path}

                # 检查是否至少选择了两个文件
                if len(selected_files) < 2:
                    messagebox.showwarning("Warning", "Please select at least two TGRT files to merge.")
                    return

                # 检查是否选择了输出文件
                output_file = output_file_var.get()
                if not output_file:
                    messagebox.showwarning("Warning", "Please select an output file.")
                    return

                # 执行合并操作
                try:
                    # 加载所有选择的文件
                    loaded_files = {}
                    student_info = None

                    for semester, file_path in selected_files.items():
                        try:
                            with open(file_path, 'rb') as f:
                                data = pickle.load(f)
                                loaded_files[semester] = data

                                # 保存第一个文件的学生信息作为参考
                                if student_info is None:
                                    student_info = data["student"]
                        except Exception as e:
                            messagebox.showerror("Error", f"Failed to load file for {semester}: {str(e)}")
                            return

                    # 检查所有文件是否属于同一个学生
                    for semester, data in loaded_files.items():
                        if data["student"]["name"] != student_info["name"] or data["student"]["student_id"] != student_info["student_id"]:
                            messagebox.showerror("Error", f"File for {semester} belongs to a different student.")
                            return

                    # 创建合并后的数据结构
                    merged_data = {
                        "student": student_info,
                        "courses": {
                            "10": {"Fall": [], "Spring": []},
                            "11": {"Fall": [], "Spring": []},
                            "12": {"Fall": [], "Spring": []}
                        }
                    }

                    # 合并所有文件中的课程数据
                    for semester, data in loaded_files.items():
                        # 解析学期信息
                        grade = semester.split()[0][1:]  # 从"G10 Fall"提取"10"
                        term = semester.split()[1]       # 从"G10 Fall"提取"Fall"

                        # 检查该学期是否有课程数据
                        if grade in data["courses"] and term in data["courses"][grade]:
                            merged_data["courses"][grade][term] = data["courses"][grade][term]

                    # 保存合并后的数据
                    with open(output_file, 'wb') as f:
                        pickle.dump(merged_data, f)

                    # 显示成功消息
                    messagebox.showinfo("Success",
                                       f"Successfully merged {len(selected_files)} files.\n"
                                       f"Merged file saved as: {output_file}")

                    # 关闭对话框
                    merge_dialog.destroy()

                except Exception as e:
                    messagebox.showerror("Error", f"Failed to merge files: {str(e)}")

            # 取消按钮回调函数
            def on_cancel():
                merge_dialog.destroy()

            # 创建按钮
            ttk.Button(btn_frame, text="Merge Files", command=on_merge).pack(side="left", padx=5)
            ttk.Button(btn_frame, text="Cancel", command=on_cancel).pack(side="right", padx=5)

            # 等待对话框关闭
            self.wait_window(merge_dialog)

        except Exception as e:
            # 如果处理过程中出现错误，显示错误消息
            messagebox.showerror("Error", f"Failed to merge TGRT files: {str(e)}")
            # 打印详细错误信息，便于调试
            import traceback
            traceback.print_exc()

    def import_excel(self):
        """
        从Excel文件导入学生成绩数据
        """
        # 1. 打开文件选择对话框
        file_path = filedialog.askopenfilename(
            filetypes=[("Excel Files", "*.xlsx;*.xls"), ("All Files", "*.*")],
            title="Select Excel File with Student Data"
        )

        # 检查用户是否取消了操作
        if not file_path:
            return  # 用户取消，直接返回

        try:
            # 2. 打开对话框让用户选择学期
            semester_dialog = tk.Toplevel(self)
            semester_dialog.title("Select Semester")
            semester_dialog.geometry("300x150")
            semester_dialog.resizable(False, False)
            semester_dialog.transient(self)
            semester_dialog.grab_set()

            # 创建学期选择框架
            frame = ttk.Frame(semester_dialog, padding=10)
            frame.pack(fill="both", expand=True)

            # 学期选择标签
            ttk.Label(frame, text="Select Semester:").pack(pady=5)

            # 学期选择下拉菜单
            semester_var = tk.StringVar()
            semester_options = [
                "G10 Fall", "G10 Spring",
                "G11 Fall", "G11 Spring",
                "G12 Fall", "G12 Spring"
            ]
            semester_combo = ttk.Combobox(
                frame,
                textvariable=semester_var,
                values=semester_options,
                state="readonly",
                width=15
            )
            semester_combo.current(0)  # 默认选择第一个选项
            semester_combo.pack(pady=5)

            # 确定和取消按钮
            btn_frame = ttk.Frame(frame)
            btn_frame.pack(fill="x", pady=10)

            # 用于存储用户选择结果的变量
            result = {"semester": None, "canceled": True}

            # 确定按钮回调函数
            def on_confirm():
                result["semester"] = semester_var.get()
                result["canceled"] = False
                semester_dialog.destroy()

            # 取消按钮回调函数
            def on_cancel():
                semester_dialog.destroy()

            # 创建按钮
            ttk.Button(btn_frame, text="Confirm", command=on_confirm).pack(side="left", padx=5)
            ttk.Button(btn_frame, text="Cancel", command=on_cancel).pack(side="right", padx=5)

            # 等待对话框关闭
            self.wait_window(semester_dialog)

            # 检查用户是否取消了操作
            if result["canceled"]:
                return

            selected_semester = result["semester"]

            # 解析选择的学期
            grade = selected_semester.split()[0].replace("G", "")  # 从"G10 Fall"提取"10"
            semester = selected_semester.split()[1]  # 从"G10 Fall"提取"Fall"

            # 3. 解析Excel文件
            # 打开Excel文件
            wb = openpyxl.load_workbook(file_path, data_only=True)
            ws = wb.active

            # 存储所有学生数据的字典
            students_data = {}

            # 从第2行开始读取数据（假设第1行是表头）
            for row in range(2, ws.max_row + 1):
                # 提取学生信息
                program = ws.cell(row=row, column=1).value  # Program
                advisor = ws.cell(row=row, column=2).value  # Advisor
                class_name = ws.cell(row=row, column=3).value  # Class Name
                class_id = ws.cell(row=row, column=4).value  # Class ID
                course = ws.cell(row=row, column=5).value  # Course
                teacher = ws.cell(row=row, column=6).value  # Teacher
                level = ws.cell(row=row, column=7).value  # Level
                student_name = ws.cell(row=row, column=8).value  # Student Name
                student_id = ws.cell(row=row, column=9).value  # Student ID
                score = ws.cell(row=row, column=10).value  # Score

                # 跳过没有学生姓名或ID的行
                if not student_name or not student_id:
                    continue

                # 将学生ID转换为字符串
                student_id = str(student_id)

                # 如果这是该学生的第一条记录，创建学生数据结构
                if student_name not in students_data:
                    students_data[student_name] = {
                        "student_id": student_id,
                        "courses": []
                    }

                # 添加课程数据
                students_data[student_name]["courses"].append({
                    "program": program,
                    "advisor": advisor,
                    "class_name": class_name,
                    "class_id": class_id,
                    "course": course,
                    "teacher": teacher,
                    "level": level,
                    "score": parse_score(score)  # 使用parse_score函数处理各种格式的分数并四舍五入为整数
                })

            # 4. 为每个学生创建.tgrt文件
            output_dir = os.path.dirname(file_path)  # 使用输入文件的目录作为输出目录
            successful_files = 0

            # 用于收集所有未匹配的课程
            all_unmatched_courses = []

            # 添加一个清理文件名的函数
            def sanitize_filename(filename):
                """
                清理文件名，移除或替换无效字符

                参数:
                    filename (str): 原始文件名

                返回:
                    str: 清理后的文件名
                """
                # Windows文件名不能包含这些字符: \ / : * ? " < > |
                invalid_chars = ['\\', '/', ':', '*', '?', '"', '<', '>', '|']

                # 替换无效字符为下划线
                for char in invalid_chars:
                    filename = filename.replace(char, '_')

                return filename

            for student_name, data in students_data.items():
                # 生成.tgrt文件名，包含学生姓名、ID和学期信息，并清理文件名中的无效字符
                sanitized_name = sanitize_filename(student_name)
                sanitized_id = sanitize_filename(str(data['student_id']))
                # 添加学期信息到文件名中
                semester_info = f"G{grade}{semester}"  # 例如：G10Fall
                tgrt_filename = f"{sanitized_name}_{sanitized_id}_{semester_info}.tgrt"
                tgrt_path = os.path.join(output_dir, tgrt_filename)

                # 打印文件名转换信息，便于调试
                if sanitized_name != student_name or sanitized_id != str(data['student_id']):
                    print(f"文件名已清理: '{student_name}_{data['student_id']}_{semester_info}.tgrt' -> '{tgrt_filename}'")

                # 确保文件路径有效
                try:
                    # 测试文件路径是否有效
                    with open(tgrt_path, 'a') as test_file:
                        pass
                    os.remove(tgrt_path)  # 删除测试文件
                except OSError as e:
                    # 如果路径无效，使用更简单的文件名
                    print(f"文件路径无效: {tgrt_path}")
                    print(f"错误信息: {str(e)}")
                    tgrt_filename = f"AP_Student_{data['student_id']}_{semester_info}.tgrt"
                    tgrt_path = os.path.join(output_dir, tgrt_filename)
                    print(f"使用替代文件名: {tgrt_filename}")

                # 检查是否存在现有的.tgrt文件
                existing_student = None
                if os.path.exists(tgrt_path):
                    try:
                        # 尝试加载现有的.tgrt文件
                        with open(tgrt_path, 'rb') as f:
                            existing_data = pickle.load(f)

                        # 创建学生对象并加载现有数据
                        existing_student = Student(
                            name=existing_data["student"].get("name", ""),
                            student_id=existing_data["student"].get("student_id", ""),
                            chinese_name=existing_data["student"].get("chinese_name", ""),
                            date_of_birth=existing_data["student"].get("date_of_birth", ""),
                            gender=existing_data["student"].get("gender", ""),
                            curriculum_program=existing_data["student"].get("curriculum_program", ""),
                            date_enrolled=existing_data["student"].get("date_enrolled", ""),
                            date_graduation=existing_data["student"].get("date_graduation", "")
                        )

                        # 加载现有的课程数据
                        for g in ["10", "11", "12"]:
                            for sem in ["Fall", "Spring"]:
                                if g in existing_data["courses"] and sem in existing_data["courses"][g]:
                                    courses = existing_data["courses"][g][sem]
                                    for i, course_info in enumerate(courses):
                                        course_key = f"{course_info['subject']}_{i+1}"
                                        existing_student.grades[g][sem][course_key] = {
                                            "subject": course_info["subject"],
                                            "course": course_info["course"],
                                            "score": course_info["score"],
                                            "scale": course_info.get("scale", "AP")
                                        }

                        print(f"已加载现有的.tgrt文件: {tgrt_path}")
                    except Exception as e:
                        print(f"加载现有.tgrt文件失败: {str(e)}")
                        existing_student = None

                # 如果存在现有的学生数据，使用它；否则创建新的学生对象
                if existing_student:
                    student = existing_student
                else:
                    student = Student(
                        name=student_name,
                        student_id=data["student_id"]
                    )

                # 获取所有预设课程名称
                all_courses = get_all_courses()

                # 用于记录未匹配的课程
                unmatched_courses = []

                # 用于跟踪每个学科的课程计数
                subject_counters = {}

                # 导入确定学科类别的函数
                from determine_subject import determine_subject_category

                # 添加课程数据
                for course_data in data["courses"]:
                    # 确定学科类别 - 使用课程实际对应的学科类别
                    original_course_name = course_data["course"]
                    # 尝试从课程名称确定学科类别
                    subject = determine_subject_category(original_course_name, GRADUATION_REQUIREMENTS)

                    # 尝试模糊匹配课程名称
                    matched_course_name = original_course_name

                    # 如果课程名称不在预设列表中，尝试模糊匹配
                    if original_course_name not in all_courses:
                        best_match, similarity, match_method, common_words = find_best_match(original_course_name, all_courses)

                        if best_match:
                            # 找到匹配的课程名称，但只有当匹配方法是"common_words"且相似度超过40%时，才使用预设课程名称
                            if match_method == "common_words" and similarity >= 0.4:
                                matched_course_name = best_match
                                use_preset_name = True
                            else:
                                # 相似度不足，使用原始课程名称
                                matched_course_name = original_course_name
                                use_preset_name = False

                            # 打印匹配信息，用于调试
                            print(f"课程匹配结果: '{original_course_name}' -> '{matched_course_name}'")
                            print(f"匹配方法: {match_method}, 相似度: {similarity:.2%}")
                            print(f"使用预设名称: {use_preset_name}")

                            # 准备匹配信息
                            match_info = ""
                            if match_method == "common_words" and common_words:
                                match_info = f"共同单词: {', '.join(common_words)}"
                                print(f"共同单词: {', '.join(common_words)}")

                            # 记录匹配信息到课程列表
                            matched_course = {
                                "student": student_name,
                                "original_course": original_course_name,
                                "matched_course": best_match,
                                "similarity": f"{similarity:.2%}",
                                "match_method": match_method,
                                "match_info": match_info,
                                "subject": subject,
                                "use_preset_name": use_preset_name,
                                "status": "matched"
                            }
                            all_unmatched_courses.append(matched_course)
                        else:
                            # 没有找到匹配的课程名称，使用原始课程名称
                            matched_course_name = original_course_name

                            # 记录到未匹配列表
                            unmatched_course = {
                                "student": student_name,
                                "original_course": original_course_name,
                                "subject": subject,
                                "status": "unmatched"
                            }
                            unmatched_courses.append(unmatched_course)
                            all_unmatched_courses.append(unmatched_course)

                    # 创建唯一的课程键
                    # 更新该学科的计数器
                    if subject not in subject_counters:
                        subject_counters[subject] = 0
                    subject_counters[subject] += 1

                    # 使用计数器创建唯一的课程键
                    course_key = f"{subject}_{subject_counters[subject]}"

                    # 将课程添加到学生的成绩数据中
                    student.grades[grade][semester][course_key] = {
                        "subject": subject,
                        "course": matched_course_name,  # 使用匹配后的课程名称
                        "score": course_data["score"],
                        "scale": "AP",  # 默认使用AP评分标准
                        "credits": 1    # 默认学分为1
                    }

                    # 打印添加的课程信息，用于调试
                    print(f"添加课程: {course_key} -> {matched_course_name} (分数: {course_data['score']})")

                # 创建要保存的数据结构
                save_data = {
                    "student": {
                        "name": student.name,
                        "student_id": student.student_id,
                        "chinese_name": student.chinese_name,
                        "date_of_birth": student.date_of_birth,
                        "gender": student.gender,
                        "curriculum_program": student.curriculum_program,
                        "date_enrolled": student.date_enrolled,
                        "date_graduation": student.date_graduation
                    },
                    "courses": {}
                }

                # 检查并修复学生对象中的小数分数问题，并应用四舍五入
                for g in ["10", "11", "12"]:
                    for sem in ["Fall", "Spring"]:
                        for course_key, course_info in student.grades[g][sem].items():
                            score = course_info["score"]
                            original_score = score

                            # 检查分数是否为小数（0-1范围内）
                            if isinstance(score, (int, float)) and 0 < score < 1:
                                # 将分数乘以100
                                score = score * 100
                                print(f"修复小数分数: {course_info['course']} - {original_score} -> {score}")

                            # 应用四舍五入
                            if isinstance(score, (int, float)) and not isinstance(score, int):
                                rounded_score = round_score(score)
                                if rounded_score != score:
                                    print(f"四舍五入分数: {course_info['course']} - {score} -> {rounded_score}")
                                score = rounded_score

                            # 更新分数
                            student.grades[g][sem][course_key]["score"] = score

                # 添加所有年级和学期的课程数据
                print(f"\n保存学生 {student_name} 的课程数据:")

                # 打印当前学期的所有课程键，用于调试
                print(f"当前学期 ({grade} {semester}) 的所有课程键:")
                for key in student.grades[grade][semester].keys():
                    print(f"  - {key}")

                for g in ["10", "11", "12"]:
                    save_data["courses"][g] = {}
                    for sem in ["Fall", "Spring"]:
                        courses = []

                        # 从学生对象中获取该学期的所有课程
                        print(f"处理学期: {g} {sem}, 课程数量: {len(student.grades[g][sem])}")
                        for course_key, course_info in student.grades[g][sem].items():
                            # 如果是当前导入的学期，打印课程信息用于调试
                            if g == grade and sem == semester:
                                print(f"  - 课程键: {course_key}")
                                print(f"    课程名: {course_info['course']}")
                                print(f"    科目: {course_info['subject']}")
                                print(f"    分数: {course_info['score']}")

                            # 添加课程数据
                            courses.append({
                                "subject": course_info["subject"],
                                "course": course_info["course"],
                                "score": course_info["score"],
                                "scale": course_info.get("scale", "AP")
                            })

                        # 保存该学期的课程列表
                        save_data["courses"][g][sem] = courses
                        print(f"已保存 {g} {sem} 学期的 {len(courses)} 门课程")

                # 保存.tgrt文件
                with open(tgrt_path, 'wb') as f:
                    pickle.dump(save_data, f)

                successful_files += 1

            # 准备显示处理结果
            result_message = f"Successfully processed {len(students_data)} students and created {successful_files} .tgrt files.\n\n"
            result_message += f"Files are saved in: {output_dir}\n\n"

            # 如果有未匹配的课程，显示报告
            if all_unmatched_courses:
                # 创建未匹配课程报告对话框
                report_dialog = tk.Toplevel(self)
                report_dialog.title("Course Matching Report")
                report_dialog.geometry("700x500")
                report_dialog.transient(self)

                # 创建报告框架
                report_frame = ttk.Frame(report_dialog, padding=10)
                report_frame.pack(fill="both", expand=True)

                # 添加说明标签
                ttk.Label(
                    report_frame,
                    text="Course Matching Report",
                    font=("Arial", 12, "bold"),
                    wraplength=680
                ).pack(pady=(0, 5))

                # 添加说明文本
                explanation_text = (
                    "This report shows how Excel course names were matched to predefined course names.\n"
                    "Matching criteria:\n"
                    "- Only courses matched by common words with similarity ≥ 40% use preset names\n"
                    "- Courses with lower similarity or matched by other methods use original names\n"
                    "- Unmatched: Course names that could not be matched to any predefined course"
                )
                ttk.Label(
                    report_frame,
                    text=explanation_text,
                    wraplength=680
                ).pack(pady=(0, 10))

                # 创建未匹配课程列表框架
                list_frame = ttk.Frame(report_frame)
                list_frame.pack(fill="both", expand=True)

                # 创建滚动条
                scrollbar = ttk.Scrollbar(list_frame)
                scrollbar.pack(side="right", fill="y")

                # 创建文本区域
                report_text = tk.Text(list_frame, wrap="word", yscrollcommand=scrollbar.set)
                report_text.pack(side="left", fill="both", expand=True)
                scrollbar.config(command=report_text.yview)

                # 配置文本标签样式
                report_text.tag_configure("title", font=("Arial", 10, "bold"))
                report_text.tag_configure("matched", foreground="green")
                report_text.tag_configure("unmatched", foreground="red")
                report_text.tag_configure("info", foreground="blue")
                report_text.tag_configure("common_words", foreground="purple")
                report_text.tag_configure("warning", foreground="orange")
                report_text.tag_configure("original_name", foreground="#8B4513")  # 棕色

                # 先添加匹配成功的课程，按是否使用预设名称分组
                matched_courses = [c for c in all_unmatched_courses if c.get("status") == "matched"]

                # 按是否使用预设名称分组
                using_preset_name = [c for c in matched_courses if c.get("use_preset_name") == True]
                using_original_name = [c for c in matched_courses if c.get("use_preset_name") == False]

                # 显示使用预设名称的匹配
                if using_preset_name:
                    report_text.insert("end", "USING PRESET COURSE NAMES (Similarity ≥ 40%):\n", "title")
                    for i, course in enumerate(using_preset_name, 1):
                        report_text.insert("end", f"{i}. Student: {course['student']}\n")
                        report_text.insert("end", f"   Original: {course['original_course']}\n")
                        report_text.insert("end", f"   Using preset name: {course['matched_course']}\n", "matched")
                        if course.get("match_info"):
                            report_text.insert("end", f"   {course['match_info']}\n", "common_words")
                        report_text.insert("end", f"   Similarity: {course['similarity']}\n", "info")
                        report_text.insert("end", f"   Subject: {course['subject']}\n\n")

                # 显示使用原始名称的匹配
                if using_original_name:
                    report_text.insert("end", "USING ORIGINAL COURSE NAMES (Similarity < 40%):\n", "title")
                    for i, course in enumerate(using_original_name, 1):
                        report_text.insert("end", f"{i}. Student: {course['student']}\n")
                        report_text.insert("end", f"   Original name (used): {course['original_course']}\n", "original_name")
                        report_text.insert("end", f"   Best match (not used): {course['matched_course']}\n")
                        report_text.insert("end", f"   Reason: Similarity below 40% threshold\n", "warning")
                        if course.get("match_info"):
                            report_text.insert("end", f"   {course['match_info']}\n", "common_words")
                        report_text.insert("end", f"   Similarity: {course['similarity']}\n", "info")
                        report_text.insert("end", f"   Subject: {course['subject']}\n\n")

                # 再添加未匹配的课程
                unmatched_courses = [c for c in all_unmatched_courses if c.get("status") == "unmatched"]
                if unmatched_courses:
                    report_text.insert("end", "UNMATCHED COURSES:\n", "title")
                    for i, course in enumerate(unmatched_courses, 1):
                        report_text.insert("end", f"{i}. Student: {course['student']}\n")
                        report_text.insert("end", f"   Course: {course['original_course']}\n", "unmatched")
                        report_text.insert("end", f"   Subject: {course['subject']}\n\n")

                # 禁用文本编辑
                report_text.config(state="disabled")

                # 添加关闭按钮
                ttk.Button(
                    report_frame,
                    text="Close",
                    command=report_dialog.destroy
                ).pack(pady=10)

                # 计算各种匹配方式的课程数量
                matched_courses = [c for c in all_unmatched_courses if c.get("status") == "matched"]
                using_preset_name_count = len([c for c in matched_courses if c.get("use_preset_name") == True])
                using_original_name_count = len([c for c in matched_courses if c.get("use_preset_name") == False])
                unmatched_count = len([c for c in all_unmatched_courses if c.get("status") == "unmatched"])

                # 在消息框中添加匹配信息
                result_message += f"\nCourse Matching Summary:\n"
                if using_preset_name_count > 0:
                    result_message += f"- {using_preset_name_count} courses are using preset names (similarity ≥ 40%)\n"
                if using_original_name_count > 0:
                    result_message += f"- {using_original_name_count} courses are using original names (similarity < 40%)\n"
                if unmatched_count > 0:
                    result_message += f"- {unmatched_count} courses could not be matched to any predefined course name\n"
                result_message += "\nSee the detailed report for more information."

            # 显示处理结果
            messagebox.showinfo("Import Completed", result_message)

        except Exception as e:
            # 如果处理过程中出现错误，显示错误消息
            messagebox.showerror("Error", f"Failed to import Excel data: {str(e)}")

    def export_pdf(self):
        """Export PDF transcript"""
        # Ensure data is up-to-date
        self.calculate_gpa()
        failed_reasons = self.check_graduation_req()

        # Collect all course data
        all_grades = {}
        for grade in ["10", "11", "12"]:
            all_grades[grade] = {}
            for semester in ["Fall", "Spring"]:
                all_grades[grade][semester] = self.get_semester_grades(grade, semester)

        # Calculate GPA for each semester
        semester_gpas = {}
        for grade in ["10", "11", "12"]:
            semester_gpas[grade] = {}
            for semester in ["Fall", "Spring"]:
                semester_gpas[grade][semester] = calculate_gpa(all_grades, grade, semester)

        # Calculate GPA for each year
        grade_gpas = {}
        for grade in ["10", "11", "12"]:
            grade_gpas[grade] = calculate_gpa(all_grades, grade)

        # Calculate overall GPA
        overall_gpa = calculate_gpa(all_grades)

        # Choose save location
        file_path = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("PDF Files", "*.pdf")],
            title="Save Transcript PDF"
        )

        if not file_path:
            return  # User canceled

        try:
            # Create PDF document
            doc = SimpleDocTemplate(file_path, pagesize=letter)
            elements = []

            # Add styles
            styles = getSampleStyleSheet()
            title_style = styles['Heading1']
            subtitle_style = styles['Heading2']
            normal_style = styles['Normal']

            # Create custom styles
            gpa_style = ParagraphStyle(
                'GPAStyle',
                parent=styles['Normal'],
                fontSize=12,
                spaceAfter=12
            )

            # Add title
            elements.append(Paragraph("Student Transcript", title_style))
            elements.append(Spacer(1, 0.1 * inch))

            # Add student information - 所有信息放在一行
            # 创建学生信息表格
            student_info_data = []

            # 添加表头行 - 包含所有字段
            header_row = ["Name in Chinese", "Name in English", "ID No", "Date of Birth", "Gender", "Curriculum Program", "Date Enrolled", "Date Graduation"]
            student_info_data.append(header_row)

            # 添加学生信息行 - 所有信息在一行
            info_row = [
                self.student.chinese_name or "",
                self.student.name or "",
                self.student.student_id or "",
                self.student.date_of_birth or "",
                self.student.gender or "",
                self.student.curriculum_program or "",
                self.student.date_enrolled or "",
                self.student.date_graduation or ""
            ]
            student_info_data.append(info_row)

            # 创建表格并设置样式 - 调整列宽和字体大小
            student_table = Table(student_info_data, colWidths=[0.8*inch, 0.9*inch, 0.6*inch, 0.7*inch, 0.5*inch, 0.8*inch, 0.7*inch, 0.7*inch])
            student_table.setStyle(TableStyle([
                # 表格边框
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                # 表头样式
                ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                # 文本对齐
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                # 字体
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 6),  # 表头字体大小更小
                ('FONTSIZE', (0, 1), (-1, 1), 6),  # 内容字体大小更小
                # 添加内边距，防止文字溢出
                ('LEFTPADDING', (0, 0), (-1, -1), 2),
                ('RIGHTPADDING', (0, 0), (-1, -1), 2),
                ('TOPPADDING', (0, 0), (-1, -1), 2),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
                # 自动换行
                ('WORDWRAP', (0, 0), (-1, -1), True),
            ]))

            elements.append(student_table)
            elements.append(Spacer(1, 0.2 * inch))

            # Add overall GPA
            elements.append(Paragraph(f"Overall GPA: {overall_gpa:.2f}", subtitle_style))
            elements.append(Spacer(1, 0.25 * inch))

            # Add each year's GPA and courses
            for grade in ["10", "11", "12"]:
                # Add year header and GPA
                if grade_gpas.get(grade, 0) > 0:
                    elements.append(Paragraph(f"Grade {grade}", subtitle_style))
                    elements.append(Paragraph(f"Year GPA: {grade_gpas[grade]:.2f}", gpa_style))

                    # Add semester GPAs
                    for semester in ["Fall", "Spring"]:
                        if semester_gpas[grade].get(semester, 0) > 0:
                            elements.append(Paragraph(
                                f"{semester} Semester GPA: {semester_gpas[grade][semester]:.2f}",
                                normal_style
                            ))

                    # Add course table
                    table_data = [['Semester', 'Course', 'Score', 'Scale', 'GPA']]

                    # Add courses from both semesters
                    for semester in ["Fall", "Spring"]:
                        courses = all_grades[grade][semester]

                        for subject, course_data in courses.items():
                            # Calculate individual course GPA
                            course_gpa = "N/A"  # Default value

                            # Check if course counts for GPA
                            scale = course_data.get("scale", "AP")

                            # Skip courses marked as "Not Included"
                            if scale == "Not Included":
                                course_gpa = "N/A"
                            else:
                                # Check if it's a non-GPA subject based on subject category
                                subject_found = False
                                subject_name = course_data.get("subject", "")  # Get the original subject name
                                for non_gpa_subject in NON_GPA_SUBJECTS:
                                    if subject_name == non_gpa_subject:  # Direct subject match
                                        subject_found = True
                                        break
                                    # Also check by course name
                                    for course_name in GRADUATION_REQUIREMENTS[non_gpa_subject]['courses']:
                                        if course_data['course'] == course_name:
                                            subject_found = True
                                            break
                                    if subject_found:
                                        break

                                if not subject_found:
                                    score = course_data["score"]

                                    # Calculate GPA
                                    for (low, high), points in GRADE_SCALE[scale].items():
                                        if low <= score <= high:
                                            course_gpa = f"{points:.1f}"
                                            break

                            table_data.append([
                                semester,
                                course_data['course'],
                                str(course_data['score']),
                                course_data.get('scale', 'AP'),
                                course_gpa
                            ])

                    if len(table_data) > 1:  # Only add table if there are courses
                        # 调整列宽以适应页面 - 优化列宽分配，确保课程名称有足够空间
                        table = Table(table_data, colWidths=[0.6*inch, 2.9*inch, 0.5*inch, 0.6*inch, 0.5*inch])

                        # 设置表格样式，调整字体大小以确保内容在表格范围内
                        table.setStyle(TableStyle([
                            # 表头样式
                            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                            ('FONTSIZE', (0, 0), (-1, 0), 8),  # 减小表头字体大小
                            ('BOTTOMPADDING', (0, 0), (-1, 0), 4),

                            # 表格内容样式
                            ('FONTSIZE', (0, 1), (-1, -1), 7),  # 减小表格内容字体大小
                            ('GRID', (0, 0), (-1, -1), 1, colors.black),
                            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),  # 垂直居中
                            ('LEFTPADDING', (0, 0), (-1, -1), 2),  # 减少左边距
                            ('RIGHTPADDING', (0, 0), (-1, -1), 2),  # 减少右边距
                            ('TOPPADDING', (0, 0), (-1, -1), 2),  # 减少上边距
                            ('BOTTOMPADDING', (0, 1), (-1, -1), 2),  # 减少下边距

                            # 课程名称列左对齐，其他列居中
                            ('ALIGN', (1, 1), (1, -1), 'LEFT'),

                            # 自动换行，确保长文本不会溢出
                            ('WORDWRAP', (0, 0), (-1, -1), True),
                        ]))
                        elements.append(table)

                    elements.append(Spacer(1, 0.5 * inch))

            # Add graduation requirements section
            elements.append(Paragraph("Graduation Requirements Status", subtitle_style))
            if not failed_reasons:
                elements.append(Paragraph("All graduation requirements have been met", gpa_style))
            else:
                elements.append(Paragraph("The following graduation requirements have not been met:", gpa_style))
                for reason in failed_reasons:
                    elements.append(Paragraph(f"• {reason}", normal_style))

            # Build PDF
            doc.build(elements)

            # Show success message
            messagebox.showinfo("Success", f"Transcript successfully exported to {file_path}")

        except Exception as e:
            messagebox.showerror("Error", f"Error exporting PDF: {str(e)}")

    def export_school_transcript(self):
        """Export School Transcript in PDF format with 9 columns layout"""
        # Ensure data is up-to-date
        self.calculate_gpa()

        # Collect all course data
        all_grades = {}
        for grade in ["10", "11", "12"]:
            all_grades[grade] = {}
            for semester in ["Fall", "Spring"]:
                all_grades[grade][semester] = self.get_semester_grades(grade, semester)

        # Calculate GPA for each year
        grade_gpas = {}
        for grade in ["10", "11", "12"]:
            grade_gpas[grade] = calculate_gpa(all_grades, grade)

        # Calculate overall GPA
        overall_gpa = calculate_gpa(all_grades)

        # Choose save location
        file_path = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("PDF Files", "*.pdf")],
            title="Save School Transcript PDF"
        )

        if not file_path:
            return  # User canceled

        try:
            # Create PDF document
            doc = SimpleDocTemplate(file_path, pagesize=letter)
            elements = []

            # Add styles
            styles = getSampleStyleSheet()
            title_style = styles['Heading1']
            subtitle_style = styles['Heading2']
            normal_style = styles['Normal']

            # 添加标题 - 确保居中
            title_style.alignment = 1  # 1表示居中对齐
            elements.append(Paragraph("Beijing National Day School Student Transcript", title_style))

            elements.append(Spacer(1, 0.1 * inch))

            # Add student information - 所有信息放在一行
            # 创建学生信息表格
            student_info_data = []

            # 添加表头行 - 包含所有字段
            header_row = ["Name in Chinese", "Name in English", "ID No", "Date of Birth", "Gender", "Curriculum Program", "Date Enrolled", "Date Graduation"]
            student_info_data.append(header_row)

            # 获取最新的学生信息 - 从UI控件获取
            student_name = self.student_name_var.get()
            student_id = self.student_id_var.get()
            chinese_name = self.student_chinese_name_var.get()
            date_of_birth = self.student_dob_var.get()
            gender = self.student_gender_var.get()
            curriculum_program = self.student_curriculum_var.get()
            date_enrolled = self.student_enrolled_var.get()
            date_graduation = self.student_graduation_var.get()

            # 更新学生对象
            self.student.name = student_name
            self.student.student_id = student_id
            self.student.chinese_name = chinese_name
            self.student.date_of_birth = date_of_birth
            self.student.gender = gender
            self.student.curriculum_program = curriculum_program
            self.student.date_enrolled = date_enrolled
            self.student.date_graduation = date_graduation

            # 打印学生信息以便调试
            print("Student Info for School Transcript:")
            print(f"Chinese Name: {chinese_name}")
            print(f"English Name: {student_name}")
            print(f"ID: {student_id}")
            print(f"DOB: {date_of_birth}")
            print(f"Gender: {gender}")
            print(f"Program: {curriculum_program}")
            print(f"Enrolled: {date_enrolled}")
            print(f"Graduation: {date_graduation}")

            # 添加学生信息行 - 确保所有信息不为空
            info_row = [
                chinese_name if chinese_name else "N/A",
                student_name if student_name else "N/A",
                student_id if student_id else "N/A",
                date_of_birth if date_of_birth else "N/A",
                gender if gender else "N/A",
                curriculum_program if curriculum_program else "N/A",
                date_enrolled if date_enrolled else "N/A",
                date_graduation if date_graduation else "N/A"
            ]
            student_info_data.append(info_row)  # 添加学生信息行

            # 创建学生信息表格 - 调整列宽以适应内容
            student_info_table = Table(student_info_data, colWidths=[0.8*inch, 0.9*inch, 0.6*inch, 0.7*inch, 0.5*inch, 0.8*inch, 0.7*inch, 0.7*inch])
            student_info_table.setStyle(TableStyle([
                # 表格边框
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                # 表头样式
                ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                # 文本对齐
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                # 字体
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 4),  # 更小的表头字体大小
                ('FONTSIZE', (0, 1), (-1, 1), 4),  # 更小的内容字体大小
                # 增加行高
                ('ROWHEIGHT', (0, 0), (-1, 1), 18),  # 增加行高，确保有足够空间显示
                # 添加内边距，防止文字溢出
                ('LEFTPADDING', (0, 0), (-1, -1), 1),  # 减少左边距
                ('RIGHTPADDING', (0, 0), (-1, -1), 1),  # 减少右边距
                ('TOPPADDING', (0, 0), (-1, -1), 1),  # 减少上边距
                ('BOTTOMPADDING', (0, 0), (-1, -1), 1),  # 减少下边距
                # 自动换行
                ('WORDWRAP', (0, 0), (-1, -1), True),
            ]))
            elements.append(student_info_table)
            elements.append(Spacer(1, 0.2 * inch))

            # 创建9列的课程表格
            # 表头: G10 Course, Fall, Spring, G11 Course, Fall, Spring, G12 Course, Fall, Spring
            course_data = []

            # 添加表头 - 只显示Fall和Spring，不显示年级编号
            header_row = [
                "G10 Course", "Fall", "Spring",
                "G11 Course", "Fall", "Spring",
                "G12 Course", "Fall", "Spring"
            ]
            course_data.append(header_row)

            # 获取每个年级的最大课程数
            max_courses = 0
            for grade in ["10", "11", "12"]:
                # 获取该年级所有课程的集合
                grade_courses = set()
                for semester in ["Fall", "Spring"]:
                    for course_id, course_info in all_grades[grade][semester].items():
                        grade_courses.add(course_info["course"])
                max_courses = max(max_courses, len(grade_courses))

            # 创建课程数据行
            # 为每个年级创建课程字典，键为课程名称，值为包含Fall和Spring分数的字典
            grade_course_dict = {}
            for grade in ["10", "11", "12"]:
                grade_course_dict[grade] = {}
                # 获取所有课程
                for semester in ["Fall", "Spring"]:
                    for course_id, course_info in all_grades[grade][semester].items():
                        course_name = course_info["course"]
                        if course_name not in grade_course_dict[grade]:
                            grade_course_dict[grade][course_name] = {"Fall": "", "Spring": ""}
                        grade_course_dict[grade][course_name][semester] = str(course_info["score"])

            # 将课程数据转换为表格行
            for i in range(max_courses):
                row = []
                for grade in ["10", "11", "12"]:
                    # 获取该年级的所有课程
                    courses = list(grade_course_dict[grade].keys())
                    if i < len(courses):
                        course_name = courses[i]
                        course_data_dict = grade_course_dict[grade][course_name]
                        row.extend([
                            course_name,
                            course_data_dict["Fall"],
                            course_data_dict["Spring"]
                        ])
                    else:
                        # 如果该年级的课程数少于最大课程数，添加空白单元格
                        row.extend(["", "", ""])
                course_data.append(row)

            # 添加GPA行
            gpa_row = []
            for grade in ["10", "11", "12"]:
                gpa_row.extend([
                    f"Year GPA: {grade_gpas[grade]:.2f}",
                    "", ""
                ])
            course_data.append(gpa_row)

            # 处理课程名称，确保长名称可以正确显示
            for i in range(1, len(course_data)-1):  # 跳过表头和GPA行
                row = course_data[i]
                # 处理每个年级的课程名称
                for col_idx in [0, 3, 6]:  # 课程名称列的索引
                    if row[col_idx] and len(row[col_idx]) > 30:  # 只有特别长的课程名称才分行（超过30个字符）
                        # 在适当位置添加换行符
                        words = row[col_idx].split()
                        if len(words) > 1:
                            # 尝试在单词之间添加换行
                            mid_point = len(words) // 2
                            first_half = " ".join(words[:mid_point])
                            second_half = " ".join(words[mid_point:])
                            row[col_idx] = f"{first_half}\n{second_half}"
                        else:
                            # 如果只有一个长单词，在中间添加换行
                            mid_point = len(row[col_idx]) // 2
                            row[col_idx] = f"{row[col_idx][:mid_point]}-\n{row[col_idx][mid_point:]}"

            # 创建课程表格 - 设置合适的列宽
            # 为课程名称列分配更多空间，为成绩列分配较少空间
            col_widths = [1.5*inch, 0.35*inch, 0.35*inch, 1.5*inch, 0.35*inch, 0.35*inch, 1.5*inch, 0.35*inch, 0.35*inch]  # 每个年级3列，共9列
            course_table = Table(course_data, colWidths=col_widths)

            # 设置表格样式
            course_table.setStyle(TableStyle([
                # 表格边框
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                # 表头背景色
                ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                # 表头字体加粗
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                # 表头字体大小
                ('FONTSIZE', (0, 0), (-1, 0), 5),  # 更小的表头字体
                # 内容字体大小
                ('FONTSIZE', (0, 1), (-1, -1), 5),  # 更小的内容字体
                # 单元格对齐方式
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                # 课程名称左对齐
                ('ALIGN', (0, 1), (0, -2), 'LEFT'),
                ('ALIGN', (3, 1), (3, -2), 'LEFT'),
                ('ALIGN', (6, 1), (6, -2), 'LEFT'),
                # GPA行背景色
                ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
                # GPA行字体加粗
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                # 合并GPA单元格
                ('SPAN', (0, -1), (2, -1)),
                ('SPAN', (3, -1), (5, -1)),
                ('SPAN', (6, -1), (8, -1)),
                # 增加行高，为多行文本提供足够空间
                ('ROWHEIGHT', (0, 1), (-1, -2), 28),  # 进一步增加课程行高度
                # 添加内边距，防止文字溢出
                ('LEFTPADDING', (0, 0), (-1, -1), 1),  # 减少内边距
                ('RIGHTPADDING', (0, 0), (-1, -1), 1),
                ('TOPPADDING', (0, 0), (-1, -1), 1),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 1),
                # 自动换行，确保长文本不会溢出
                ('WORDWRAP', (0, 0), (-1, -1), True),
            ]))
            elements.append(course_table)

            # 添加GPA计算方法说明
            elements.append(Spacer(1, 0.2 * inch))

            # 创建GPA计算方法表格 - 使用更易理解的表达形式
            gpa_info_data = [
                ["GPA Calculation Methods"],
                ["AP Scale: A+ (97-100) = 4.3 points, A (90-96) = 4.0 points, B (80-89) = 3.5 points,"],
                ["                C (70-79) = 3.0 points, D (60-69) = 2.0 points, F (0-59) = 0.0 points"],
                ["CNCC Scale: A+ (97-100) = 4.3 points, A (85-96) = 4.0 points, B (70-84) = 3.0 points,"],
                ["                  C (60-69) = 2.0 points, F (0-59) = 0.0 points"],
                [f"Overall GPA: {overall_gpa:.2f}"]
            ]

            # 创建GPA信息表格
            gpa_info_table = Table(gpa_info_data, colWidths=[6.7*inch])
            gpa_info_table.setStyle(TableStyle([
                # 表格边框
                ('BOX', (0, 0), (-1, -1), 1, colors.black),
                # 标题背景色
                ('BACKGROUND', (0, 0), (0, 0), colors.lightgrey),
                # 标题字体加粗
                ('FONTNAME', (0, 0), (0, 0), 'Helvetica-Bold'),
                # 字体大小
                ('FONTSIZE', (0, 0), (-1, -1), 6),  # 小字体
                # Overall GPA字体加粗
                ('FONTNAME', (0, -1), (0, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, -1), (0, -1), 8),  # Overall GPA字体稍大
                # 单元格对齐方式
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),  # 标题居中
                ('ALIGN', (0, 1), (-1, 4), 'LEFT'),    # 计算方法左对齐
                ('ALIGN', (0, -1), (-1, -1), 'CENTER'), # Overall GPA居中
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                # 内边距
                ('LEFTPADDING', (0, 0), (-1, -1), 4),
                ('RIGHTPADDING', (0, 0), (-1, -1), 4),
                ('TOPPADDING', (0, 0), (-1, -1), 2),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
                # 自动换行
                ('WORDWRAP', (0, 0), (-1, -1), True),
            ]))
            elements.append(gpa_info_table)

            # 添加空间
            elements.append(Spacer(1, 0.5 * inch))

            # 添加签名和盖章行
            # 获取当前日期
            from datetime import datetime
            current_date = datetime.now().strftime("%Y-%m-%d")

            # 创建签名表格数据
            signature_data = [
                ["Signature of Principal", "Date Transcript Issued", "Seal of the School"],
                ["_________________", current_date, "_________________"]
            ]

            # 创建签名表格
            signature_table = Table(signature_data, colWidths=[2.2*inch, 2.2*inch, 2.2*inch])
            signature_table.setStyle(TableStyle([
                # 表格边框
                ('BOX', (0, 0), (-1, -1), 1, colors.black),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                # 表头样式
                ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                # 表头字体加粗
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                # 字体大小
                ('FONTSIZE', (0, 0), (-1, 0), 8),  # 表头字体
                ('FONTSIZE', (0, 1), (-1, 1), 8),  # 内容字体
                # 单元格对齐方式
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                # 行高
                ('ROWHEIGHT', (0, 1), (-1, 1), 30),  # 为签名留出足够空间
                # 内边距
                ('LEFTPADDING', (0, 0), (-1, -1), 4),
                ('RIGHTPADDING', (0, 0), (-1, -1), 4),
                ('TOPPADDING', (0, 0), (-1, -1), 2),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
            ]))
            elements.append(signature_table)

            # 添加空间
            elements.append(Spacer(1, 0.5 * inch))

            # 创建学校信息样式
            school_info_style = ParagraphStyle(
                'SchoolInfoStyle',
                parent=styles['Normal'],
                fontSize=8,
                alignment=0,  # 左对齐
                leading=10    # 行间距
            )

            # 添加学校信息
            elements.append(Paragraph("66 yuquan Road, Haidian District, Beijing, China 100039", school_info_style))
            elements.append(Paragraph("Tel:0086-10-88625495, Fax:0086-10-88278424", school_info_style))
            elements.append(Paragraph("Http:// www.bnds.cn", school_info_style))

            # Build PDF
            doc.build(elements)

            # Show success message
            messagebox.showinfo("Success", f"School Transcript successfully exported to {file_path}")

        except Exception as e:
            messagebox.showerror("Error", f"Error exporting School Transcript: {str(e)}")

    def update_student_data(self):
        """
        更新学生数据，包括学生基本信息和所有已完成的课程

        1. 从UI获取学生所有信息
        2. 创建新的学生对象
        3. 收集所有年级和学期的课程数据
        4. 更新毕业要求完成情况
        """
        # 1. 从UI获取学生所有信息
        student_name = self.student_name_var.get()
        student_id = self.student_id_var.get()
        chinese_name = self.student_chinese_name_var.get()
        date_of_birth = self.student_dob_var.get()
        gender = self.student_gender_var.get()
        curriculum_program = self.student_curriculum_var.get()
        date_enrolled = self.student_enrolled_var.get()
        date_graduation = self.student_graduation_var.get()

        # 2. 创建新的学生对象，使用最新的信息
        self.student = Student(
            name=student_name,
            student_id=student_id,
            chinese_name=chinese_name,
            date_of_birth=date_of_birth,
            gender=gender,
            curriculum_program=curriculum_program,
            date_enrolled=date_enrolled,
            date_graduation=date_graduation
        )

        # 3. 收集所有年级和学期的课程数据
        for grade in ["10", "11", "12"]:
            for semester in ["Fall", "Spring"]:
                # 获取该学期的所有课程成绩
                grades = self.get_semester_grades(grade, semester)
                # 将课程数据存储到学生对象中
                self.student.grades[grade][semester] = grades

                # 4. 更新毕业要求完成情况
                for _, data in grades.items():
                    course_name = data["course"]
                    # 查找课程所属的学科类别
                    for req_subject, req_data in GRADUATION_REQUIREMENTS.items():
                        # 检查课程是否属于该学科
                        if course_name in req_data['courses']:
                            # 特殊处理中国社会科学课程 - 需要记录具体完成的课程
                            if req_subject == 'Chinese_Social_Studies':
                                # 将课程添加到已完成课程集合中
                                self.student.requirements[req_subject]["taken_courses"].add(course_name)
                            else:
                                # 对于其他学科，增加已完成的学期数
                                self.student.requirements[req_subject]["taken"] += 1
                            # 找到匹配的学科后跳出循环
                            break

class GradeFrame(ttk.Frame):
    """
    年级框架类

    管理特定年级(10/11/12)的课程数据，包含秋季和春季两个学期的标签页。
    提供复制秋季课程到春季学期的功能。

    属性:
        grade (str): 年级 ("10", "11", "12")
        course_db (dict): 课程数据库
        semester_notebook (ttk.Notebook): 学期标签页
    """
    def __init__(self, parent, grade, course_db):
        """
        初始化年级框架

        参数:
            parent: 父级窗口组件
            grade (str): 年级 ("10", "11", "12")
            course_db (dict): 课程数据库
        """
        # 初始化基类
        super().__init__(parent)

        # 保存年级信息和课程数据库
        self.grade = grade
        self.course_db = course_db

        # ===== 创建控制按钮区域 =====
        self.control_frame = ttk.Frame(self)
        self.control_frame.pack(fill="x", padx=5, pady=5)

        # 添加"复制秋季到春季"按钮
        self.copy_btn = ttk.Button(
            self.control_frame,
            text="Copy Fall to Spring",
            command=self.copy_fall_to_spring  # 点击时调用复制方法
        )
        self.copy_btn.pack(side="left", padx=5)

        # ===== 创建学期选择标签页 =====
        self.semester_notebook = ttk.Notebook(self)
        # 为秋季和春季学期创建标签页
        for semester in ["Fall", "Spring"]:
            # 创建学期框架
            sem_frame = SemesterFrame(self.semester_notebook, grade, semester, course_db)
            # 添加到标签页
            self.semester_notebook.add(sem_frame, text=semester)
        # 将标签页添加到年级框架
        self.semester_notebook.pack(expand=True, fill="both")

    def copy_fall_to_spring(self):
        """
        将秋季学期的课程复制到春季学期

        1. 获取秋季和春季学期的框架
        2. 获取秋季学期的所有课程
        3. 清空春季学期的现有课程
        4. 将秋季学期的课程添加到春季学期
        """
        # 1. 获取秋季和春季学期的框架
        fall_frame = None
        spring_frame = None

        # 遍历所有学期框架
        for sem_frame in self.semester_notebook.winfo_children():
            if isinstance(sem_frame, SemesterFrame):
                # 找到秋季学期框架
                if sem_frame.semester == "Fall":
                    fall_frame = sem_frame
                # 找到春季学期框架
                elif sem_frame.semester == "Spring":
                    spring_frame = sem_frame

        # 检查是否找到了两个学期框架
        if not fall_frame or not spring_frame:
            messagebox.showwarning("Warning", "Could not find semester frames")
            return

        # 2. 获取秋季学期的所有课程
        fall_courses = fall_frame.get_courses()
        # 检查秋季学期是否有课程
        if not fall_courses:
            messagebox.showinfo("Information", "No courses in Fall semester to copy")
            return

        # 3. 清空春季学期的现有课程
        # 删除所有课程项
        for item_id in spring_frame.courses_list.get_children():
            spring_frame.courses_list.delete(item_id)
        # 重置课程条目列表
        spring_frame.course_entries = []

        # 4. 将秋季学期的课程添加到春季学期
        for subject, course, score, scale in fall_courses:
            # 确保分数是整数
            rounded_score = round_score(score)

            # 添加到课程列表
            item_id = spring_frame.courses_list.insert("", "end", values=(subject, course, rounded_score, scale))
            # 添加到课程条目列表
            spring_frame.course_entries.append((subject, course, rounded_score, scale, item_id))

        # 显示复制成功消息
        messagebox.showinfo("Success", f"Copied {len(fall_courses)} courses from Fall to Spring semester")

class SemesterFrame(ttk.Frame):
    """
    学期框架类

    管理特定年级特定学期的课程数据，提供添加、编辑、删除课程的功能。

    属性:
        grade (str): 年级 ("10", "11", "12")
        semester (str): 学期 ("Fall", "Spring")
        course_db (dict): 课程数据库
        course_entries (list): 保存所有课程条目的列表
    """
    def __init__(self, parent, grade, semester, course_db):
        """
        初始化学期框架

        参数:
            parent: 父级窗口组件
            grade (str): 年级 ("10", "11", "12")
            semester (str): 学期 ("Fall", "Spring")
            course_db (dict): 课程数据库
        """
        # 初始化基类
        super().__init__(parent)

        # 保存基本信息
        self.grade = grade  # 年级信息
        self.semester = semester  # 学期信息
        self.course_db = course_db  # 课程数据库
        self.course_entries = []  # 保存所有课程条目的列表

        # ===== 创建主布局 =====
        self.main_frame = ttk.Frame(self)
        self.main_frame.pack(fill="both", expand=True, padx=10, pady=10)

        # ===== 创建添加课程区域 =====
        self.add_frame = ttk.LabelFrame(self.main_frame, text="Add New Course")
        self.add_frame.pack(fill="x", padx=5, pady=5)

        # 学科下拉菜单
        ttk.Label(self.add_frame, text="Subject:").grid(row=0, column=0, padx=5, pady=5)
        self.subject_var = tk.StringVar()  # 存储选择的学科
        self.subject_combo = ttk.Combobox(
            self.add_frame,
            textvariable=self.subject_var,
            values=list(course_db.keys())  # 使用课程数据库中的所有学科作为选项
        )
        self.subject_combo.grid(row=0, column=1, padx=5, pady=5)
        # 绑定选择事件，当选择学科时更新课程下拉菜单
        self.subject_combo.bind("<<ComboboxSelected>>", self.update_courses)

        # 课程下拉菜单
        ttk.Label(self.add_frame, text="Course:").grid(row=0, column=2, padx=5, pady=5)
        self.course_var = tk.StringVar()  # 存储选择的课程
        self.course_combo = ttk.Combobox(self.add_frame, textvariable=self.course_var, width=20)
        self.course_combo.grid(row=0, column=3, padx=5, pady=5)

        # 分数输入
        ttk.Label(self.add_frame, text="Score:").grid(row=0, column=4, padx=5, pady=5)
        self.score_var = tk.IntVar(value=90)  # 默认分数为90
        self.score_spin = ttk.Spinbox(
            self.add_frame,
            from_=0, to=100,  # 分数范围0-100
            textvariable=self.score_var,
            width=5
        )
        self.score_spin.grid(row=0, column=5, padx=5, pady=5)

        # 评分标准选择
        ttk.Label(self.add_frame, text="Scale:").grid(row=0, column=6, padx=5, pady=5)
        self.scale_var = tk.StringVar(value="AP")  # 默认使用AP评分标准
        self.scale_combo = ttk.Combobox(
            self.add_frame,
            textvariable=self.scale_var,
            values=["AP", "CNCC", "Not Included"],  # 评分标准选项
            width=12
        )
        self.scale_combo.grid(row=0, column=7, padx=5, pady=5)

        # 添加按钮
        self.add_btn = ttk.Button(
            self.add_frame,
            text="Add Course",
            command=self.add_course  # 点击时调用添加课程方法
        )
        self.add_btn.grid(row=0, column=8, padx=5, pady=5)

        # ===== 创建课程列表区域 =====
        self.courses_frame = ttk.LabelFrame(self.main_frame, text="Added Courses")
        self.courses_frame.pack(fill="both", expand=True, padx=5, pady=5)

        # 创建课程列表(树形视图)
        self.courses_list = ttk.Treeview(
            self.courses_frame,
            columns=("subject", "course", "score", "scale"),  # 列名
            show="headings"  # 只显示数据列，不显示树形结构
        )
        # 设置列标题
        self.courses_list.heading("subject", text="Subject")
        self.courses_list.heading("course", text="Course")
        self.courses_list.heading("score", text="Score")
        self.courses_list.heading("scale", text="Scale")
        # 设置列宽
        self.courses_list.column("subject", width=100)
        self.courses_list.column("course", width=200)
        self.courses_list.column("score", width=50)
        self.courses_list.column("scale", width=80)
        # 添加到框架
        self.courses_list.pack(fill="both", expand=True, padx=5, pady=5)

        # ===== 创建按钮区域 =====
        self.buttons_frame = ttk.Frame(self.courses_frame)
        self.buttons_frame.pack(fill="x", pady=5)

        # 编辑按钮
        self.edit_btn = ttk.Button(
            self.buttons_frame,
            text="Edit Selected Course",
            command=self.edit_course  # 点击时调用编辑课程方法
        )
        self.edit_btn.pack(side="left", padx=5)

        # 删除按钮
        self.delete_btn = ttk.Button(
            self.buttons_frame,
            text="Delete Selected Course",
            command=self.delete_course  # 点击时调用删除课程方法
        )
        self.delete_btn.pack(side="left", padx=5)

    def update_courses(self, _=None):
        """
        根据选择的学科更新课程下拉菜单

        当用户在学科下拉菜单中选择一个学科时，更新课程下拉菜单显示该学科下的所有课程。

        参数:
            _: Tkinter事件对象，由ComboBox的<<ComboboxSelected>>事件触发，不使用
        """
        # 获取选择的学科
        subject = self.subject_var.get()
        # 检查学科是否在课程数据库中
        if subject in self.course_db:
            # 更新课程下拉菜单的选项
            self.course_combo["values"] = self.course_db[subject]
            # 如果该学科有课程，默认选择第一个
            if self.course_db[subject]:
                self.course_combo.current(0)

    def add_course(self):
        """
        添加新课程到列表

        从输入字段获取课程信息，验证后添加到课程列表中，并清空输入字段。
        """
        # 获取输入的课程信息
        subject = self.subject_var.get()  # 学科
        course = self.course_var.get()    # 课程名称
        raw_score = self.score_var.get()  # 原始分数
        scale = self.scale_var.get()      # 评分标准

        # 将分数四舍五入为整数
        score = round_score(raw_score)

        # 验证必填字段
        if not subject or not course:
            # 如果学科或课程为空，显示警告
            tk.messagebox.showwarning("Input Error", "Please select a subject and course")
            return

        # 添加到课程列表(树形视图)
        item_id = self.courses_list.insert("", "end", values=(subject, course, score, scale))
        # 添加到课程条目列表，包括树形视图中的项目ID
        self.course_entries.append((subject, course, score, scale, item_id))

        # 清空输入字段，准备下一次输入
        self.subject_combo.set("")  # 清空学科选择
        self.course_combo.set("")   # 清空课程选择
        self.score_var.set(90)      # 重置分数为默认值90
        self.scale_var.set("AP")    # 重置评分标准为默认值AP

    def edit_course(self):
        """
        编辑选中的课程

        打开编辑对话框，允许用户修改选中课程的分数和评分标准。
        学科和课程名称不可修改，只能修改分数和评分标准。
        """
        # 获取选中的课程项
        selected = self.courses_list.selection()
        # 验证是否只选中了一个课程
        if not selected or len(selected) != 1:
            messagebox.showinfo("Information", "Please select exactly one course to edit")
            return

        # 获取选中项的ID
        item_id = selected[0]
        # 查找对应的课程条目
        course_entry = None
        for entry in self.course_entries:
            if entry[4] == item_id:
                course_entry = entry
                break

        # 如果找不到对应的课程条目，直接返回
        if not course_entry:
            return

        # ===== 创建编辑对话框 =====
        edit_dialog = tk.Toplevel(self)
        edit_dialog.title("Edit Course")
        edit_dialog.geometry("400x300")  # 设置对话框大小
        edit_dialog.resizable(False, False)  # 禁止调整大小
        edit_dialog.transient(self)  # 设置为模态对话框
        edit_dialog.grab_set()  # 获取焦点

        # ===== 可编辑的课程信息区域 =====
        info_frame = ttk.LabelFrame(edit_dialog, text="Course Information")
        info_frame.pack(fill="x", padx=10, pady=10)

        # 学科(可修改)
        ttk.Label(info_frame, text="Subject:").grid(row=0, column=0, padx=5, pady=5, sticky="w")
        subject_var = tk.StringVar(value=course_entry[0])
        # 获取所有可能的学科类别
        all_subjects = list(GRADUATION_REQUIREMENTS.keys())
        subject_combo = ttk.Combobox(info_frame, textvariable=subject_var, values=all_subjects, width=20)
        subject_combo.grid(row=0, column=1, padx=5, pady=5, sticky="w")

        # 课程名称(可修改)
        ttk.Label(info_frame, text="Course:").grid(row=1, column=0, padx=5, pady=5, sticky="w")
        course_var = tk.StringVar(value=course_entry[1])
        course_entry_widget = ttk.Entry(info_frame, textvariable=course_var, width=30)
        course_entry_widget.grid(row=1, column=1, padx=5, pady=5, sticky="w")

        # ===== 可编辑字段区域 =====
        edit_frame = ttk.LabelFrame(edit_dialog, text="Edit Fields")
        edit_frame.pack(fill="x", padx=10, pady=10)

        # 分数输入
        ttk.Label(edit_frame, text="Score:").grid(row=0, column=0, padx=5, pady=5, sticky="w")
        score_var = tk.IntVar(value=course_entry[2])  # 使用当前分数作为默认值
        score_spin = ttk.Spinbox(edit_frame, from_=0, to=100, textvariable=score_var, width=5)
        score_spin.grid(row=0, column=1, padx=5, pady=5, sticky="w")

        # 评分标准选择
        ttk.Label(edit_frame, text="Scale:").grid(row=1, column=0, padx=5, pady=5, sticky="w")
        scale_var = tk.StringVar(value=course_entry[3])  # 使用当前评分标准作为默认值
        scale_combo = ttk.Combobox(edit_frame, textvariable=scale_var, values=["AP", "CNCC", "Not Included"], width=12)
        scale_combo.grid(row=1, column=1, padx=5, pady=5, sticky="w")

        # 添加分隔线，提高视觉清晰度
        ttk.Separator(edit_dialog, orient="horizontal").pack(fill="x", padx=10, pady=5)

        # ===== 按钮区域 =====
        button_frame = ttk.Frame(edit_dialog)
        button_frame.pack(fill="x", padx=10, pady=10)

        # 保存更改的函数
        def save_changes():
            # 获取修改后的值
            _, _, old_score, old_scale, entry_id = course_entry
            new_subject = subject_var.get()
            new_course = course_var.get()
            raw_score = score_var.get()
            new_scale = scale_var.get()

            # 验证必填字段
            if not new_subject or not new_course:
                messagebox.showwarning("Input Error", "Subject and Course cannot be empty")
                return

            # 将分数四舍五入为整数
            new_score = round_score(raw_score)

            # 更新树形视图中的显示
            self.courses_list.item(item_id, values=(new_subject, new_course, new_score, new_scale))

            # 更新内部数据
            index = self.course_entries.index(course_entry)
            self.course_entries[index] = (new_subject, new_course, new_score, new_scale, entry_id)

            # 关闭对话框
            edit_dialog.destroy()
            # 显示成功消息
            messagebox.showinfo("Success", "Course updated successfully")

        # 创建保存按钮
        save_button = ttk.Button(button_frame, text="Save Changes", command=save_changes, width=15)
        save_button.pack(side="left", padx=10, pady=5)

        # 创建取消按钮
        cancel_button = ttk.Button(button_frame, text="Cancel", command=edit_dialog.destroy, width=10)
        cancel_button.pack(side="right", padx=10, pady=5)

    def delete_course(self):
        """
        删除选中的课程

        从课程列表和内部数据中删除选中的课程。
        可以同时删除多个选中的课程。
        """
        # 获取选中的课程项
        selected = self.courses_list.selection()
        # 如果没有选中任何课程，直接返回
        if not selected:
            return

        # 从列表和内部数据中删除选中的课程
        for item_id in selected:
            # 从树形视图中删除
            self.courses_list.delete(item_id)
            # 从内部数据中删除
            self.course_entries = [entry for entry in self.course_entries if entry[4] != item_id]

    def get_courses(self):
        """
        获取所有已添加的课程数据

        返回:
            list: 课程数据列表，每个元素为(学科, 课程名称, 分数, 评分标准)的元组
        """
        # 返回所有课程数据，不包括树形视图中的项目ID
        return [(entry[0], entry[1], entry[2], entry[3]) for entry in self.course_entries]

# ================== 主程序入口 ==================
if __name__ == "__main__":
    """
    程序入口点

    创建并启动GradeTracker应用程序。
    包含全局异常处理，确保任何未捕获的异常都会显示给用户。
    """
    try:
        # 创建GradeTracker应用实例
        app = GradeTracker()
        # 启动主事件循环
        app.mainloop()
    except Exception as e:
        # 捕获并显示任何未处理的异常
        messagebox.showerror("Error", f"Program error: {str(e)}")
        # 同时在控制台打印详细错误信息，便于调试
        print(f"Error details: {e}")
